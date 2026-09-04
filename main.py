"""
CRM Association "Les Ami(e)s de Romy"
Backend FastAPI - multi-utilisateurs, audit trail, emailing, événements, documents
Déploiement: Raspberry Pi (Tailscale) + Windows local
"""

import json
import os
import re
import base64
import smtplib
import time
import threading
import hashlib
import secrets
import urllib.request
import urllib.error
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Request, Response, HTTPException, Depends, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ============ CONFIG ============
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "db.json"
BACKUP_DIR = BASE_DIR / "backups"
AUTH_PATH = BASE_DIR / "auth_config.json"
SMTP_PATH = BASE_DIR / "smtp_config.json"
DOCS_DIR = BASE_DIR / "uploads"
STORAGE_DIR = BASE_DIR / "nipogi_storage"
STORAGE_UPLOADS = STORAGE_DIR / "uploads"
STORAGE_DOCUMENTS = STORAGE_DIR / "documents"
STORAGE_BACKUPS = STORAGE_DIR / "backups"
STORAGE_ARCHIVES = STORAGE_DIR / "archives"
BACKUP_DIR.mkdir(exist_ok=True)
DOCS_DIR.mkdir(exist_ok=True)
# Nipogi storage via SSHFS mount — create subdirs if mount is present
for _d in (STORAGE_DIR, STORAGE_UPLOADS, STORAGE_DOCUMENTS, STORAGE_BACKUPS, STORAGE_ARCHIVES):
    try:
        _d.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass  # mount not available — endpoints will report gracefully

VERSION = "1.34"
HOST = os.environ.get("CRM_HOST", "100.89.45.97")
PORT = int(os.environ.get("CRM_PORT", "8001"))

# ============ AUTH ============
def load_auth():
    with open(AUTH_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def load_smtp():
    try:
        with open(SMTP_PATH, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        # Support base64-encoded password (security: not in plaintext)
        if "password_b64" in cfg and not cfg.get("password"):
            cfg["password"] = base64.b64decode(cfg["password_b64"]).decode("utf-8")
        return cfg
    except Exception:
        return {"password": ""}

def get_user_from_token(token: str):
    auth = load_auth()
    for u in auth["users"]:
        if u["token"] == token:
            return u
    return None

def require_auth(request: Request):
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token requis")
    token = auth_header[7:]
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token invalide")
    return user

def require_admin(user=Depends(require_auth)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces admin requis")
    return user

REFERENT_ROLES = {"admin", "referent"}

def require_referent(user=Depends(require_auth)):
    """Referent or admin only. Members get 403."""
    if user["role"] not in REFERENT_ROLES:
        raise HTTPException(status_code=403, detail="Acces referent requis")
    return user

def is_member(user):
    return user["role"] == "membre"

def is_referent_or_admin(user):
    return user["role"] in REFERENT_ROLES

# ============ PASSWORD HASHING (pbkdf2) ============
def hash_password(password: str, salt: str = None):
    if salt is None:
        salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    return key.hex(), salt

def verify_password(password: str, stored_hash: str, salt: str) -> bool:
    key, _ = hash_password(password, salt)
    return secrets.compare_digest(key, stored_hash)

def migrate_plaintext_passwords():
    """Auto-hash plaintext passwords on startup."""
    auth = load_auth()
    changed = False
    for u in auth["users"]:
        if "password" in u and "password_hash" not in u:
            pwd = u.pop("password")
            h, s = hash_password(pwd)
            u["password_hash"] = h
            u["password_salt"] = s
            changed = True
    if changed:
        with open(AUTH_PATH, "w", encoding="utf-8") as f:
            json.dump(auth, f, ensure_ascii=False, indent=2)
        print("[SECURITY] Mots de passe migrés vers hash pbkdf2")

# ============ RATE LIMITING (login) ============
_login_attempts = {}
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW = 300  # 5 minutes

def check_login_rate_limit(client_ip: str):
    now = time.time()
    if client_ip in _login_attempts:
        _login_attempts[client_ip] = [t for t in _login_attempts[client_ip] if now - t < _LOGIN_WINDOW]
        if len(_login_attempts[client_ip]) >= _LOGIN_MAX_ATTEMPTS:
            raise HTTPException(status_code=429, detail="Trop de tentatives. Reessayez dans 5 minutes.")
    else:
        _login_attempts[client_ip] = []

def record_failed_login(client_ip: str):
    if client_ip not in _login_attempts:
        _login_attempts[client_ip] = []
    _login_attempts[client_ip].append(time.time())

def clear_login_attempts(client_ip: str):
    _login_attempts.pop(client_ip, None)

# ============ DB (atomic write + recovery) ============
_db_lock = threading.Lock()

def load_db():
    try:
        with open(DB_PATH, "r", encoding="utf-8") as f:
            return json.loads(f.read(), strict=False)
    except json.JSONDecodeError:
        # Try recovery from backup
        latest = _find_latest_valid_backup()
        if latest:
            with open(latest, "r", encoding="utf-8") as f:
                return json.loads(f.read(), strict=False)
        return {"contacts": [], "evenements": [], "documents": []}
    except FileNotFoundError:
        return {"contacts": [], "evenements": [], "documents": []}

def save_db(data):
    with _db_lock:
        tmp = DB_PATH.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        try:
            os.replace(tmp, DB_PATH)
        except OSError:
            # Fallback for bind-mounted files (os.replace fails with "Device or resource busy")
            with open(DB_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            try:
                os.unlink(tmp)
            except OSError:
                pass

def backup_db():
    with _db_lock:
        if not DB_PATH.exists():
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bk = BACKUP_DIR / f"db_backup_{ts}.json"
        tmp = bk.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(load_db(), f, ensure_ascii=False, indent=2)
        os.replace(tmp, bk)
        # Keep only last 30 backups
        backups = sorted(BACKUP_DIR.glob("db_backup_*.json"))
        for old in backups[:-30]:
            old.unlink(missing_ok=True)

def _find_latest_valid_backup():
    backups = sorted(BACKUP_DIR.glob("db_backup_*.json"), reverse=True)
    for bk in backups:
        try:
            with open(bk, "r", encoding="utf-8") as f:
                d = json.loads(f.read(), strict=False)
            if d.get("contacts"):
                return bk
        except Exception:
            continue
    return None

def validate_db_integrity():
    data = load_db()
    if not data.get("contacts") and not data.get("evenements") and not data.get("documents"):
        latest = _find_latest_valid_backup()
        if latest:
            with open(latest, "r", encoding="utf-8") as f:
                data = json.loads(f.read(), strict=False)
            save_db(data)
            print(f"[RECOVERY] Base restaurée depuis {latest.name}")
    return data

# ============ MODELS ============
class LoginRequest(BaseModel):
    username: str
    password: str

class ContactUpdate(BaseModel):
    prenom: str = ""
    nom: str = ""
    qualite: str = ""
    telephone: str = ""
    email: str = ""
    notes: str = ""
    tags: str = ""  # tags separes par virgules

class ContactCreate(ContactUpdate):
    pass

class HistoriqueEntry(BaseModel):
    action: str
    details: str = ""

class EvenementCreate(BaseModel):
    titre: str
    date: str
    heure: str = ""
    lieu: str = ""
    description: str = ""
    participants: list[str] = []
    type: str = ""  # "", "intervention" (sensibilisation), etc.

class EvenementUpdate(BaseModel):
    titre: Optional[str] = None
    date: Optional[str] = None
    heure: Optional[str] = None
    lieu: Optional[str] = None
    description: Optional[str] = None
    participants: Optional[list[str]] = None
    presence: Optional[dict] = None  # {contact_id: "present"|"absent"|"excuse"|""}
    type: Optional[str] = None

class CotisationCreate(BaseModel):
    contact_id: str
    annee: str  # ex "2026"
    montant: float = 0.0
    date_paiement: str = ""  # YYYY-MM-DD
    mode_paiement: str = ""  # especes, cheque, virement
    statut: str = "paye"  # paye, en_attente, impaye
    notes: str = ""

class CotisationUpdate(BaseModel):
    montant: Optional[float] = None
    date_paiement: Optional[str] = None
    mode_paiement: Optional[str] = None
    statut: Optional[str] = None
    notes: Optional[str] = None

class TacheCreate(BaseModel):
    titre: str
    description: str = ""
    assigne_a: str = ""  # contact_id
    echeance: str = ""  # YYYY-MM-DD
    priorite: str = "normale"  # basse, normale, haute, urgente

class TacheUpdate(BaseModel):
    titre: Optional[str] = None
    description: Optional[str] = None
    assigne_a: Optional[str] = None
    echeance: Optional[str] = None
    priorite: Optional[str] = None
    statut: Optional[str] = None  # a_faire, en_cours, terminee

class DonCreate(BaseModel):
    contact_id: str = ""
    nom: str = ""
    montant: float = 0.0
    date: str = ""
    mode_paiement: str = ""
    type_don: str = "don"  # don, sponsor, subvention
    notes: str = ""

class FeedbackCreate(BaseModel):
    type: str = "bug"  # bug, suggestion, autre
    message: str
    page: str = ""  # sur quelle page/view

class EmailSend(BaseModel):
    subject: str
    body: str
    recipients: list[str] = []
    contact_ids: list[str] = []
    template: str = ""
    html: bool = False

class DocumentCreate(BaseModel):
    nom: str
    type_doc: str = ""
    description: str = ""

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "editeur"
    contact_id: str = ""  # for membre role: link to their contact record

class ChangePassword(BaseModel):
    current_password: str
    new_password: str

class UserUpdate(BaseModel):
    password: Optional[str] = None
    role: Optional[str] = None
    contact_id: Optional[str] = None

class PresenceSelf(BaseModel):
    """Member self-registration for an event."""
    contact_id: str
    presence: str = "present"  # present, absent, excuse, ""

# ============ MODELS V1.26 — nouvelles fonctions ============
class AccompagnementCreate(BaseModel):
    contact_id: str
    type_suivi: str = "accompagnement"  # accompagnement, suivi famille, orientation, signalement
    date_debut: str = ""  # YYYY-MM-DD
    date_fin: str = ""
    statut: str = "actif"  # actif, clos, suspendu
    priorite: str = "normale"  # basse, normale, haute, urgente
    intervenants: list[str] = []  # contact_ids des référents
    notes_confidentielles: str = ""
    notes_partagees: str = ""

class AccompagnementUpdate(BaseModel):
    type_suivi: Optional[str] = None
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None
    statut: Optional[str] = None
    priorite: Optional[str] = None
    intervenants: Optional[list[str]] = None
    notes_confidentielles: Optional[str] = None
    notes_partagees: Optional[str] = None

class AccompagnementNote(BaseModel):
    note: str
    confidentiel: bool = False

class BenevolePlanning(BaseModel):
    evenement_id: str
    contact_id: str
    role: str = "participant"  # participant, responsable, animateur, logistique
    creneau: str = ""  # ex: "matin", "apres-midi", "09:00-12:00"

# ============ MODELS V1.27 — gestion quotidienne ============
class NoteFraisCreate(BaseModel):
    contact_id: str = ""  # qui a engage la depense
    date: str = ""  # YYYY-MM-DD
    montant: float = 0.0
    categorie: str = ""  # deplacement, repas, fournitures, communication, autre
    description: str = ""
    justificatif: str = ""  # nom du fichier upload
    statut: str = "en_attente"  # en_attente, valide, rembourse

class NoteFraisUpdate(BaseModel):
    montant: Optional[float] = None
    categorie: Optional[str] = None
    description: Optional[str] = None
    statut: Optional[str] = None

class EcheanceCreate(BaseModel):
    titre: str
    date: str = ""  # YYYY-MM-DD
    type_echeance: str = "administratif"  # administratif, fiscal, assurance, ag, reunion, autre
    description: str = ""
    recursif: str = ""  # annuel, mensuel, "" (vide = ponctuel)
    responsable: str = ""

class EcheanceUpdate(BaseModel):
    titre: Optional[str] = None
    date: Optional[str] = None
    type_echeance: Optional[str] = None
    description: Optional[str] = None
    recursif: Optional[str] = None
    responsable: Optional[str] = None
    statut: Optional[str] = None  # a_venir, fait, en_retard

class RegistreAppelCreate(BaseModel):
    contact_id: str = ""  # contact appele (si dans la base)
    nom_appelant: str = ""
    nom_appele: str = ""
    date_appel: str = ""  # YYYY-MM-DD
    heure: str = ""
    motif: str = ""  # information, demande aide, signalement, suivi, autre
    description: str = ""
    suite_donnee: str = ""

class VoteCreate(BaseModel):
    titre: str
    date: str = ""  # YYYY-MM-DD
    type_vote: str = "ag"  # ag, ca, bureau
    description: str = ""

class VoteResultat(BaseModel):
    pour: int = 0
    contre: int = 0
    abstention: int = 0
    notes: str = ""

class ChecklistCreate(BaseModel):
    titre: str
    type_checklist: str = "evenement"  # evenement, cloture_exercice, ag, import_excel, autre
    taches: list[str] = []  # liste des items a cocher
    evenement_id: str = ""

class ChecklistUpdate(BaseModel):
    taches: Optional[list[str]] = None
    coche: Optional[list[bool]] = None

# ============ MODELS V1.28 — communication & gestion ============
class SondageCreate(BaseModel):
    question: str
    options: list[str] = []  # choix possibles
    cible: str = "membres"  # membres, referents, tous
    date_fin: str = ""  # YYYY-MM-DD
    multiple: bool = False  # reponses multiples autorisees

class SondageReponse(BaseModel):
    option_index: int  # index dans options

class CompteRenduCreate(BaseModel):
    titre: str
    date: str = ""  # YYYY-MM-DD
    type_reunion: str = "reunion"  # reunion, ag, ca, bureau
    presents: list[str] = []  # contact_ids
    absents: list[str] = []
    excused: list[str] = []
    ordre_du_jour: str = ""
    discussions: str = ""
    decisions: str = ""
    actions: str = ""  # actions a suivre

class CompteRenduUpdate(BaseModel):
    titre: Optional[str] = None
    date: Optional[str] = None
    type_reunion: Optional[str] = None
    presents: Optional[list[str]] = None
    absents: Optional[list[str]] = None
    excused: Optional[list[str]] = None
    ordre_du_jour: Optional[str] = None
    discussions: Optional[str] = None
    decisions: Optional[str] = None
    actions: Optional[str] = None

class SmsCampaignCreate(BaseModel):
    message: str
    destinataires: list[str] = []  # numeros de telephone
    contact_ids: list[str] = []  # ou par contact_id
    qualite: str = ""  # ou par groupe qualite

class PresseContactCreate(BaseModel):
    nom: str
    media: str = ""  # nom du media
    fonction: str = ""  # journaliste, redacteur, etc.
    email: str = ""
    telephone: str = ""
    notes: str = ""

class PresseReleaseCreate(BaseModel):
    sujet: str
    date: str = ""
    media_cible: str = ""  # nom du media ou "tous"
    contacts_envoyes: list[str] = []  # presse_contact_ids
    contenu: str = ""
    statut: str = "brouillon"  # brouillon, envoye, publie

class PresseCouvertureCreate(BaseModel):
    titre: str
    media: str = ""
    date: str = ""
    type_couverture: str = "article"  # article, emission, interview, mention
    lien: str = ""
    resume: str = ""

class MembreProfilUpdate(BaseModel):
    telephone: Optional[str] = None
    email: Optional[str] = None
    ville: Optional[str] = None
    code_postal: Optional[str] = None

# ============ MODELS V1.34 — subventions ============
class SubventionCreate(BaseModel):
    organisme: str  # organisme financeur (ex: Conseil departemental, Fonds de dotation)
    intitule: str  # nom de l'appel a projets / dispositif
    montant_demande: float = 0.0
    montant_accorde: float = 0.0
    date_demande: str = ""  # YYYY-MM-DD
    date_reponse: str = ""  # YYYY-MM-DD
    statut: str = "brouillon"  # brouillon, depose, accepte, refuse
    echeance: str = ""  # YYYY-MM-DD date limite de depot
    documents_requis: str = ""  # documents a fournir (texte libre)
    documents_remis: str = ""  # documents effectivement remis (texte libre)
    notes: str = ""

class SubventionUpdate(BaseModel):
    organisme: Optional[str] = None
    intitule: Optional[str] = None
    montant_demande: Optional[float] = None
    montant_accorde: Optional[float] = None
    date_demande: Optional[str] = None
    date_reponse: Optional[str] = None
    statut: Optional[str] = None
    echeance: Optional[str] = None
    documents_requis: Optional[str] = None
    documents_remis: Optional[str] = None
    notes: Optional[str] = None

# ============ APP ============
app = FastAPI(title="CRM Les Ami(e)s de Romy")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://100.89.45.97:8001", "http://localhost:8001", "http://127.0.0.1:8001"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

# ============ SECURITY HEADERS MIDDLEWARE ============
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    """Add security headers to all responses."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https:; "
        "font-src 'self'; "
        "connect-src 'self'; "
        "frame-ancestors 'none'"
    )
    return response

# ============ HELPERS ============
def now_iso():
    return datetime.now().isoformat()

def normalize_phone(val):
    if not val:
        return ""
    digits = re.sub(r'\D', '', str(val))
    if len(digits) == 9 and digits[0] in '67':
        digits = '0' + digits
    if len(digits) == 10:
        return f"{digits[0:2]} {digits[2:4]} {digits[4:6]} {digits[6:8]} {digits[8:10]}"
    return str(val).strip()

def add_history(contact, user, action, details=""):
    entry = {
        "date": now_iso(),
        "user": user["username"],
        "action": action,
        "details": details
    }
    contact.setdefault("historique", []).append(entry)
    contact["modifie_par"] = user["username"]
    contact["modifie_le"] = entry["date"]

def gen_id(prefix="romy"):
    data = load_db()
    existing = {c["id"] for c in data["contacts"]}
    n = len(data["contacts"]) + 1
    while f"{prefix}-{n:03d}" in existing:
        n += 1
    return f"{prefix}-{n:03d}"

def send_email_smtp(to_emails, subject, body, attachments=None, html=False):
    smtp = load_smtp()
    if not smtp.get("password"):
        raise HTTPException(status_code=400, detail="SMTP non configure (smtp_config.json)")
    
    msg = MIMEMultipart()
    msg["From"] = "vincentminvielle@gmail.com"
    msg["To"] = ", ".join(to_emails)
    msg["Subject"] = subject
    
    if html:
        msg.attach(MIMEText(body, "html", "utf-8"))
    else:
        msg.attach(MIMEText(body, "plain", "utf-8"))
    
    if attachments:
        for filepath in attachments:
            if not os.path.exists(filepath):
                continue
            with open(filepath, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(filepath)}"')
                msg.attach(part)
    
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login("vincentminvielle@gmail.com", smtp["password"])
        server.sendmail("vincentminvielle@gmail.com", to_emails, msg.as_string())

# ============ AUTH ENDPOINTS ============
@app.post("/auth/login")
async def login(req: LoginRequest, request: Request):
    if not isinstance(req.username, str) or not isinstance(req.password, str):
        raise HTTPException(status_code=422, detail="Types invalides")
    client_ip = request.client.host if request.client else "unknown"
    check_login_rate_limit(client_ip)
    auth = load_auth()
    for u in auth["users"]:
        if u["username"] == req.username:
            stored_hash = u.get("password_hash")
            stored_salt = u.get("password_salt")
            if stored_hash and stored_salt:
                if verify_password(req.password, stored_hash, stored_salt):
                    clear_login_attempts(client_ip)
                    return {"status": "ok", "username": u["username"], "role": u["role"], "token": u["token"], "contact_id": u.get("contact_id", "")}
            elif u.get("password") == req.password:
                # Legacy plaintext - migrate on successful login
                h, s = hash_password(req.password)
                u.pop("password", None)
                u["password_hash"] = h
                u["password_salt"] = s
                with open(AUTH_PATH, "w", encoding="utf-8") as f:
                    json.dump(auth, f, ensure_ascii=False, indent=2)
                clear_login_attempts(client_ip)
                return {"status": "ok", "username": u["username"], "role": u["role"], "token": u["token"], "contact_id": u.get("contact_id", "")}
            record_failed_login(client_ip)
            raise HTTPException(status_code=401, detail="Identifiants invalides")
    record_failed_login(client_ip)
    raise HTTPException(status_code=401, detail="Identifiants invalides")

@app.get("/auth/verify")
async def verify_token(token: str = Query(...)):
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token invalide")
    return {"status": "ok", "username": user["username"], "role": user["role"], "contact_id": user.get("contact_id", "")}

@app.get("/auth/users")
async def list_users(user=Depends(require_admin)):
    auth = load_auth()
    return [{"username": u["username"], "role": u["role"], "contact_id": u.get("contact_id", "")} for u in auth["users"]]

@app.post("/auth/users")
async def create_user(req: UserCreate, user=Depends(require_admin)):
    auth = load_auth()
    if any(u["username"] == req.username for u in auth["users"]):
        raise HTTPException(status_code=409, detail="Utilisateur existe deja")
    import secrets
    token = f"romy-{req.role}-" + secrets.token_hex(8)
    pwd_hash, pwd_salt = hash_password(req.password)
    new_user = {"username": req.username, "password_hash": pwd_hash, "password_salt": pwd_salt, "role": req.role, "token": token}
    if req.contact_id:
        new_user["contact_id"] = req.contact_id
    auth["users"].append(new_user)
    with open(AUTH_PATH, "w", encoding="utf-8") as f:
        json.dump(auth, f, ensure_ascii=False, indent=2)
    return {"status": "ok", "username": req.username, "role": req.role}

@app.put("/auth/users/{username}")
async def update_user(username: str, req: UserUpdate, user=Depends(require_admin)):
    auth = load_auth()
    for u in auth["users"]:
        if u["username"] == username:
            if req.password:
                h, s = hash_password(req.password)
                u["password_hash"] = h
                u["password_salt"] = s
                u.pop("password", None)
            if req.role:
                u["role"] = req.role
            if req.contact_id is not None:
                u["contact_id"] = req.contact_id
            with open(AUTH_PATH, "w", encoding="utf-8") as f:
                json.dump(auth, f, ensure_ascii=False, indent=2)
            return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Utilisateur non trouve")

@app.post("/auth/change-password")
async def change_password(req: ChangePassword, user=Depends(require_auth)):
    """Permet a tout utilisateur connecte de changer son propre mot de passe."""
    auth = load_auth()
    for u in auth["users"]:
        if u["username"] == user["username"]:
            # Verifier l'ancien mot de passe
            stored_hash = u.get("password_hash")
            stored_salt = u.get("password_salt")
            if stored_hash and stored_salt:
                if not verify_password(req.current_password, stored_hash, stored_salt):
                    raise HTTPException(status_code=403, detail="Mot de passe actuel incorrect")
            elif u.get("password") != req.current_password:
                raise HTTPException(status_code=403, detail="Mot de passe actuel incorrect")
            # Valider le nouveau mot de passe
            if len(req.new_password) < 4:
                raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit faire au moins 4 caracteres")
            # Hasher et sauvegarder
            h, s = hash_password(req.new_password)
            u["password_hash"] = h
            u["password_salt"] = s
            u.pop("password", None)
            with open(AUTH_PATH, "w", encoding="utf-8") as f:
                json.dump(auth, f, ensure_ascii=False, indent=2)
            return {"status": "ok", "message": "Mot de passe modifie avec succes"}
    raise HTTPException(status_code=404, detail="Utilisateur non trouve")

@app.delete("/auth/users/{username}")
async def delete_user(username: str, user=Depends(require_admin)):
    if username == "vincent":
        raise HTTPException(status_code=403, detail="Impossible de supprimer l'admin principal")
    auth = load_auth()
    auth["users"] = [u for u in auth["users"] if u["username"] != username]
    with open(AUTH_PATH, "w", encoding="utf-8") as f:
        json.dump(auth, f, ensure_ascii=False, indent=2)
    return {"status": "ok"}

# ============ CONTACTS ENDPOINTS ============
@app.get("/api/contacts")
async def list_contacts(q: str = "", qualite: str = "", user=Depends(require_auth)):
    data = load_db()
    contacts = data["contacts"]
    # Members: only see other members and referents (not donors, journalists, etc.)
    if is_member(user):
        contacts = [c for c in contacts if c.get("qualite", "").lower() in ("membre asso", "referent", "benevole")]
    if q:
        ql = q.lower()
        contacts = [c for c in contacts if ql in (c.get("prenom","") + " " + c.get("nom","") + " " + c.get("email","") + " " + c.get("telephone","") + " " + c.get("notes","") + " " + c.get("tags","") + " " + c.get("qualite","")).lower()]
    if qualite:
        contacts = [c for c in contacts if c.get("qualite","").lower() == qualite.lower()]
    return contacts

@app.get("/api/contacts/{contact_id}")
async def get_contact(contact_id: str, user=Depends(require_auth)):
    data = load_db()
    for c in data["contacts"]:
        if c["id"] == contact_id:
            # Members: cannot view contacts outside member/referent/benevole
            if is_member(user):
                if c.get("qualite", "").lower() not in ("membre asso", "referent", "benevole"):
                    raise HTTPException(status_code=403, detail="Acces refuse")
            return c
    raise HTTPException(status_code=404, detail="Contact non trouve")

@app.post("/api/contacts")
async def create_contact(req: ContactCreate, user=Depends(require_referent)):
    data = load_db()
    contact = {
        "id": gen_id(),
        "prenom": req.prenom.strip(),
        "nom": req.nom.strip().upper(),
        "qualite": req.qualite.strip(),
        "telephone": normalize_phone(req.telephone),
        "email": req.email.strip().lower(),
        "notes": req.notes.strip(),
        "tags": req.tags.strip(),
        "historique": [],
        "cree_par": user["username"],
        "cree_le": now_iso(),
        "modifie_par": "",
        "modifie_le": ""
    }
    add_history(contact, user, "creation", f"Contact cree par {user['username']}")
    data["contacts"].append(contact)
    save_db(data)
    backup_db()
    return {"status": "ok", "contact": contact}

@app.put("/api/contacts/{contact_id}")
async def update_contact(contact_id: str, req: ContactUpdate, user=Depends(require_referent)):
    data = load_db()
    for c in data["contacts"]:
        if c["id"] == contact_id:
            changes = []
            if c["prenom"] != req.prenom:
                changes.append(f"prenom: '{c['prenom']}' -> '{req.prenom}'")
                c["prenom"] = req.prenom.strip()
            if c["nom"].upper() != req.nom.strip().upper():
                changes.append(f"nom: '{c['nom']}' -> '{req.nom}'")
                c["nom"] = req.nom.strip().upper()
            if c["qualite"] != req.qualite:
                changes.append(f"qualite: '{c['qualite']}' -> '{req.qualite}'")
                c["qualite"] = req.qualite.strip()
            old_tel = c["telephone"]
            new_tel = normalize_phone(req.telephone)
            if old_tel != new_tel:
                changes.append(f"telephone: '{old_tel}' -> '{new_tel}'")
                c["telephone"] = new_tel
            if c["email"] != req.email.lower():
                changes.append(f"email: '{c['email']}' -> '{req.email}'")
                c["email"] = req.email.strip().lower()
            if c["notes"] != req.notes:
                changes.append(f"notes modifiees")
                c["notes"] = req.notes.strip()
            if c.get("tags", "") != req.tags:
                changes.append(f"tags: '{c.get('tags', '')}' -> '{req.tags}'")
                c["tags"] = req.tags.strip()
            if changes:
                add_history(c, user, "modification", " | ".join(changes))
                save_db(data)
                backup_db()
            return {"status": "ok", "contact": c}
    raise HTTPException(status_code=404, detail="Contact non trouve")

@app.delete("/api/contacts/{contact_id}")
async def delete_contact(contact_id: str, user=Depends(require_referent)):
    data = load_db()
    before = len(data["contacts"])
    contact = None
    for c in data["contacts"]:
        if c["id"] == contact_id:
            contact = c
            break
    if not contact:
        raise HTTPException(status_code=404, detail="Contact non trouve")
    data["contacts"] = [c for c in data["contacts"] if c["id"] != contact_id]
    save_db(data)
    backup_db()
    return {"status": "ok", "deleted": contact_id}

@app.post("/api/contacts/{contact_id}/historique")
async def add_historique(contact_id: str, req: HistoriqueEntry, user=Depends(require_referent)):
    data = load_db()
    for c in data["contacts"]:
        if c["id"] == contact_id:
            add_history(c, user, req.action, req.details)
            save_db(data)
            backup_db()
            return {"status": "ok", "contact": c}
    raise HTTPException(status_code=404, detail="Contact non trouve")

@app.get("/api/contacts/{contact_id}/historique")
async def get_historique(contact_id: str, user=Depends(require_referent)):
    data = load_db()
    for c in data["contacts"]:
        if c["id"] == contact_id:
            return c.get("historique", [])
    raise HTTPException(status_code=404, detail="Contact non trouve")

# ============ EVENEMENTS ENDPOINTS ============

def _format_date_fr(date_str):
    """Convertit YYYY-MM-DD en date francaise lisible, ex: mercredi 15 juillet 2026."""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        jours = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
        mois = ["janvier", "fevrier", "mars", "avril", "mai", "juin",
                "juillet", "aout", "septembre", "octobre", "novembre", "decembre"]
        return f"{jours[dt.weekday()]} {dt.day} {mois[dt.month - 1]} {dt.year}"
    except Exception:
        return date_str

def _send_event_notification(ev, data, user):
    """
    Envoie un email automatique aux participants d'un evenement qui ont un email.
    Retourne un dict {sent_to, count} si un email a ete envoye, None sinon.
    Si SMTP n'est pas configure ou aucun participant n'a d'email, retourne None (silencieux).
    """
    # Verifier que SMTP est configure
    smtp = load_smtp()
    if not smtp.get("password"):
        return None

    # Recuperer les contacts participants
    participant_ids = ev.get("participants", [])
    if not participant_ids:
        return None

    contacts_by_id = {c["id"]: c for c in data["contacts"]}
    participant_contacts = []
    for pid in participant_ids:
        c = contacts_by_id.get(pid)
        if c and c.get("email") and c["email"].strip():
            participant_contacts.append(c)

    if not participant_contacts:
        return None

    # Construire la liste des noms des participants (tous, meme sans email)
    all_names = []
    for pid in participant_ids:
        c = contacts_by_id.get(pid)
        if c:
            all_names.append(f"{c.get('prenom', '')} {c.get('nom', '')}".strip())
        else:
            all_names.append(pid)
    participants_str = ", ".join(all_names)

    # Construire le sujet et le corps
    sujet = f"Evenement : {ev['titre']}"

    lignes = []
    lignes.append(f"Bonjour,")
    lignes.append("")
    lignes.append(f"Un evenement a ete cree dans le CRM de l'association Les Ami(e)s de Romy :")
    lignes.append("")
    lignes.append(f"Titre : {ev['titre']}")
    date_fr = _format_date_fr(ev.get("date", ""))
    if ev.get("heure"):
        lignes.append(f"Date : {date_fr} a {ev['heure']}")
    else:
        lignes.append(f"Date : {date_fr}")
    if ev.get("lieu"):
        lignes.append(f"Lieu : {ev['lieu']}")
    if ev.get("description"):
        lignes.append(f"Description : {ev['description']}")
    lignes.append(f"Participants : {participants_str}")
    lignes.append(f"Cree par : {user['username']}")
    lignes.append("")
    lignes.append("Cordialement,")
    lignes.append("L'equipe des Ami(e)s de Romy")

    body = "\n".join(lignes)

    # Recuperer les emails des participants qui en ont un
    recipient_emails = [c["email"].strip() for c in participant_contacts]

    try:
        send_email_smtp(recipient_emails, sujet, body)
    except Exception:
        # Si l'envoi echoue, on ne bloque pas la creation de l'evenement
        return None

    # Ajouter une entree d'historique a chaque participant notifie
    for c in participant_contacts:
        add_history(c, user, "email", f"Notification evenement: {ev['titre']}")

    return {"sent_to": recipient_emails, "count": len(recipient_emails)}

@app.get("/api/evenements")
async def list_evenements(user=Depends(require_auth)):
    data = load_db()
    return data.get("evenements", [])

@app.post("/api/evenements")
async def create_evenement(req: EvenementCreate, user=Depends(require_referent)):
    data = load_db()
    ev = {
        "id": f"evt-{secrets.token_hex(8)}",
        "titre": req.titre,
        "date": req.date,
        "heure": req.heure,
        "lieu": req.lieu,
        "description": req.description,
        "participants": req.participants,
        "type": req.type,
        "presence": {},
        "rappel_envoye": False,
        "cree_par": user["username"],
        "cree_le": now_iso(),
        "historique": [{"date": now_iso(), "user": user["username"], "action": "creation"}]
    }
    data["evenements"].append(ev)
    save_db(data)
    backup_db()

    # --- Envoi email automatique aux participants qui ont un email ---
    email_info = _send_event_notification(ev, data, user)
    if email_info:
        ev["email_sent"] = email_info
        save_db(data)

    return {"status": "ok", "evenement": ev}

@app.put("/api/evenements/{evt_id}")
async def update_evenement(evt_id: str, req: EvenementUpdate, user=Depends(require_referent)):
    data = load_db()
    for ev in data["evenements"]:
        if ev["id"] == evt_id:
            changes = []
            if req.titre is not None and ev["titre"] != req.titre:
                changes.append(f"titre modifie")
                ev["titre"] = req.titre
            if req.date is not None and ev["date"] != req.date:
                changes.append(f"date: '{ev['date']}' -> '{req.date}'")
                ev["date"] = req.date
            if req.heure is not None and ev.get("heure", "") != req.heure:
                changes.append(f"heure: '{ev.get('heure', '')}' -> '{req.heure}'")
                ev["heure"] = req.heure
            if req.lieu is not None and ev["lieu"] != req.lieu:
                changes.append(f"lieu modifie")
                ev["lieu"] = req.lieu
            if req.description is not None and ev["description"] != req.description:
                changes.append("description modifiee")
                ev["description"] = req.description
            if req.participants is not None:
                old = set(ev.get("participants", []))
                new = set(req.participants)
                added = new - old
                removed = old - new
                if added:
                    changes.append(f"participants ajoutes: {', '.join(added)}")
                if removed:
                    changes.append(f"participants retires: {', '.join(removed)}")
                ev["participants"] = req.participants
            if req.presence is not None:
                ev["presence"] = req.presence
                changes.append("presence mise a jour")
            if req.type is not None and ev.get("type", "") != req.type:
                changes.append(f"type: '{ev.get('type', '')}' -> '{req.type}'")
                ev["type"] = req.type
            if changes:
                ev.setdefault("historique", []).append({"date": now_iso(), "user": user["username"], "action": "modification", "details": " | ".join(changes)})
            save_db(data)
            backup_db()
            return {"status": "ok", "evenement": ev}
    raise HTTPException(status_code=404, detail="Evenement non trouve")

@app.delete("/api/evenements/{evt_id}")
async def delete_evenement(evt_id: str, user=Depends(require_referent)):
    data = load_db()
    data["evenements"] = [e for e in data["evenements"] if e["id"] != evt_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}

# ============ INTERVENTIONS (sensibilisation par etablissement) ============
@app.get("/api/interventions/synthese")
async def interventions_synthese(user=Depends(require_auth)):
    """Synthese des evenements de type 'intervention' groupes par etablissement (lieu).
    Retourne: liste triee par nb interventions desc, avec pour chaque etablissement:
    - nom, nb_interventions, dates[], nb_total_participants, derniere_date, frequence
    """
    data = load_db()
    evenements = data.get("evenements", [])

    # Filtrer les evenements de type intervention
    interventions = [ev for ev in evenements if ev.get("type", "") == "intervention"]

    # Grouper par etablissement (lieu)
    par_lieu = {}
    for ev in interventions:
        lieu = ev.get("lieu", "").strip() or "(Non precise)"
        if lieu not in par_lieu:
            par_lieu[lieu] = {
                "etablissement": lieu,
                "nb_interventions": 0,
                "dates": [],
                "nb_total_participants": 0,
            }
        grp = par_lieu[lieu]
        grp["nb_interventions"] += 1
        ev_date = ev.get("date", "")
        if ev_date:
            grp["dates"].append(ev_date)
        nb_part = len(ev.get("participants", []))
        grp["nb_total_participants"] += nb_part

    # Construire la liste de synthese
    synthese = []
    for lieu, grp in par_lieu.items():
        dates_triees = sorted(grp["dates"], reverse=True)
        derniere_date = dates_triees[0] if dates_triees else ""
        # Frequence: intervalle moyen en jours entre interventions
        frequence = ""
        if len(dates_triees) >= 2:
            from datetime import datetime as _dt
            dates_asc = sorted(grp["dates"])
            intervals = []
            for i in range(1, len(dates_asc)):
                try:
                    d1 = _dt.strptime(dates_asc[i - 1], "%Y-%m-%d")
                    d2 = _dt.strptime(dates_asc[i], "%Y-%m-%d")
                    intervals.append((d2 - d1).days)
                except Exception:
                    pass
            if intervals:
                avg_days = sum(intervals) / len(intervals)
                if avg_days < 7:
                    frequence = f"Tous les {round(avg_days)} j"
                elif avg_days < 30:
                    frequence = f"Toutes les {round(avg_days / 7)} sem"
                elif avg_days < 365:
                    frequence = f"Tous les {round(avg_days / 30)} mois"
                else:
                    frequence = f"Tous les {round(avg_days / 365)} an(s)"
            else:
                frequence = "Ponctuel"
        else:
            frequence = "Ponctuel"

        synthese.append({
            "etablissement": grp["etablissement"],
            "nb_interventions": grp["nb_interventions"],
            "dates": dates_triees,
            "nb_total_participants": grp["nb_total_participants"],
            "derniere_date": derniere_date,
            "frequence": frequence,
        })

    # Trier par nb interventions desc
    synthese.sort(key=lambda x: x["nb_interventions"], reverse=True)

    total_interventions = sum(s["nb_interventions"] for s in synthese)
    total_participants = sum(s["nb_total_participants"] for s in synthese)
    nb_etablissements = len(synthese)

    return {
        "total_interventions": total_interventions,
        "total_participants": total_participants,
        "nb_etablissements": nb_etablissements,
        "synthese": synthese,
    }

# ============ PRESENCE (membre self-registration) ============
@app.post("/api/evenements/{evt_id}/presence")
async def set_self_presence(evt_id: str, req: PresenceSelf, user=Depends(require_auth)):
    """Any logged-in user (including members) can set their own presence."""
    data = load_db()
    for ev in data["evenements"]:
        if ev["id"] == evt_id:
            presence = ev.get("presence", {})
            presence[req.contact_id] = req.presence
            ev["presence"] = presence
            ev.setdefault("historique", []).append({
                "date": now_iso(),
                "user": user["username"],
                "action": "presence_auto",
                "details": f"Contact {req.contact_id}: {req.presence}"
            })
            save_db(data)
            backup_db()
            return {"status": "ok", "evenement": ev}
    raise HTTPException(status_code=404, detail="Evenement non trouve")

# ============ COTISATIONS ENDPOINTS ============
@app.get("/api/cotisations")
async def list_cotisations(annee: str = "", contact_id: str = "", user=Depends(require_referent)):
    data = load_db()
    cotisations = data.get("cotisations", [])
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    if annee:
        cotisations = [c for c in cotisations if c.get("annee") == annee]
    if contact_id:
        cotisations = [c for c in cotisations if c.get("contact_id") == contact_id]
    # Enrichir avec le nom du contact
    for cot in cotisations:
        c = contact_by_id.get(cot.get("contact_id"), {})
        cot["contact_nom"] = f"{c.get('prenom','')} {c.get('nom','')}".strip()
        cot["contact_email"] = c.get("email", "")
    return cotisations

@app.post("/api/cotisations")
async def create_cotisation(req: CotisationCreate, user=Depends(require_referent)):
    data = load_db()
    cot = {
        "id": f"cot-{secrets.token_hex(8)}",
        "contact_id": req.contact_id,
        "annee": req.annee,
        "montant": req.montant,
        "date_paiement": req.date_paiement,
        "mode_paiement": req.mode_paiement,
        "statut": req.statut,
        "notes": req.notes,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data.setdefault("cotisations", []).append(cot)
    for c in data["contacts"]:
        if c["id"] == req.contact_id:
            add_history(c, user, "cotisation", f"Annee {req.annee} | {req.statut} | {req.montant} EUR")
            break
    save_db(data)
    backup_db()
    return {"status": "ok", "cotisation": cot}

@app.put("/api/cotisations/{cot_id}")
async def update_cotisation(cot_id: str, req: CotisationUpdate, user=Depends(require_referent)):
    data = load_db()
    for cot in data.get("cotisations", []):
        if cot["id"] == cot_id:
            if req.montant is not None:
                cot["montant"] = req.montant
            if req.date_paiement is not None:
                cot["date_paiement"] = req.date_paiement
            if req.mode_paiement is not None:
                cot["mode_paiement"] = req.mode_paiement
            if req.statut is not None:
                cot["statut"] = req.statut
            if req.notes is not None:
                cot["notes"] = req.notes
            save_db(data)
            backup_db()
            return {"status": "ok", "cotisation": cot}
    raise HTTPException(status_code=404, detail="Cotisation non trouvee")

@app.delete("/api/cotisations/{cot_id}")
async def delete_cotisation(cot_id: str, user=Depends(require_referent)):
    data = load_db()
    data["cotisations"] = [c for c in data.get("cotisations", []) if c["id"] != cot_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}

# ============ RAPPEL EMAIL AUTOMATIQUE (J-2) ============
def _send_event_reminder(ev, data):
    smtp = load_smtp()
    if not smtp.get("password"):
        return False
    participant_ids = ev.get("participants", [])
    if not participant_ids:
        return False
    contacts_by_id = {c["id"]: c for c in data["contacts"]}
    recipient_emails = []
    for pid in participant_ids:
        c = contacts_by_id.get(pid)
        if c and c.get("email") and c["email"].strip():
            recipient_emails.append(c["email"].strip())
    if not recipient_emails:
        return False
    sujet = f"Rappel : {ev['titre']}"
    date_fr = _format_date_fr(ev.get("date", ""))
    lignes = [
        "Bonjour,",
        "",
        "Ceci est un rappel pour l evenement suivant qui aura lieu prochainement :",
        "",
        f"Titre : {ev['titre']}",
        f"Date : {date_fr}" + (f" a {ev['heure']}" if ev.get("heure") else ""),
        f"Lieu : {ev.get('lieu', 'non precise')}" if ev.get("lieu") else "",
        "",
        "Au plaisir de vous y retrouver,",
        "L equipe des Ami(e)s de Romy"
    ]
    body = "\n".join(lignes)
    try:
        send_email_smtp(recipient_emails, sujet, body)
        return True
    except Exception:
        return False

def _reminder_check_loop():
    """Thread de fond: scrute toutes les 30 min les evenements a J-2."""
    while True:
        time.sleep(1800)
        try:
            data = load_db()
            smtp = load_smtp()
            if not smtp.get("password"):
                continue
            today = datetime.now().date()
            for ev in data.get("evenements", []):
                if ev.get("rappel_envoye"):
                    continue
                if not ev.get("date"):
                    continue
                try:
                    ev_date = datetime.strptime(ev["date"], "%Y-%m-%d").date()
                except Exception:
                    continue
                delta = (ev_date - today).days
                if delta == 2:
                    sent = _send_event_reminder(ev, data)
                    if sent:
                        ev["rappel_envoye"] = True
                        save_db(data)
                        print(f"[RAPPEL] Rappel envoye pour: {ev['titre']}")
        except Exception as e:
            print(f"[RAPPEL] Erreur: {e}")

# ============ DOCUMENTS ENDPOINTS ============
@app.get("/api/documents")
async def list_documents(user=Depends(require_referent)):
    data = load_db()
    return data.get("documents", [])

@app.post("/api/documents")
async def upload_document(request: Request, user=Depends(require_referent)):
    form = await request.form()
    nom = form.get("nom", "")
    type_doc = form.get("type_doc", "")
    description = form.get("description", "")
    file = form.get("file")
    
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier requis")
    
    # restriction des extensions autorisees
    ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".txt", ".csv", ".odt", ".ods", ".odp", ".rtf", ".zip"}
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Type de fichier non autorise: {ext}")
    
    # limite de taille: 20 Mo
    MAX_FILE_SIZE = 20 * 1024 * 1024
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 20 Mo)")
    
    safe_name = re.sub(r'[^\w\.-]', '_', file.filename)
    # eviter path traversal
    safe_name = safe_name.replace("..", "").replace("/", "").replace("\\", "")
    filepath = DOCS_DIR / safe_name
    with open(filepath, "wb") as f:
        f.write(content)
    
    data = load_db()
    doc = {
        "id": f"doc-{secrets.token_hex(8)}",
        "nom": nom or safe_name,
        "type_doc": type_doc,
        "description": description,
        "filename": safe_name,
        "filepath": str(filepath),
        "taille": len(content),
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data.setdefault("documents", []).append(doc)
    save_db(data)
    return {"status": "ok", "document": doc}

@app.get("/api/documents/{doc_id}")
async def download_document(doc_id: str, user=Depends(require_referent)):
    data = load_db()
    for d in data.get("documents", []):
        if d["id"] == doc_id:
            # Chercher le fichier dans uploads/ (ancien) ou nipogi_storage/documents/ (nouveau)
            filepath = d.get("filepath", "")
            if not filepath and d.get("fichier"):
                # Nouveau format: fichier = "asso/nom.pdf" dans nipogi_storage/documents/
                filepath = str(STORAGE_DIR / "documents" / d["fichier"])
                if not os.path.exists(filepath):
                    # Fallback: uploads/
                    filepath = str(DOCS_DIR / d["fichier"])
            if not os.path.exists(filepath):
                raise HTTPException(status_code=404, detail="Fichier introuvable")
            return FileResponse(filepath, filename=d.get("filename", d.get("nom", "document")))
    raise HTTPException(status_code=404, detail="Document non trouve")

@app.delete("/api/documents/{doc_id}")
async def delete_document(doc_id: str, user=Depends(require_referent)):
    data = load_db()
    for d in data.get("documents", []):
        if d["id"] == doc_id:
            # Supprimer le fichier (ancien ou nouveau format)
            filepath = d.get("filepath", "")
            if not filepath and d.get("fichier"):
                filepath = str(STORAGE_DIR / "documents" / d["fichier"])
                if not os.path.exists(filepath):
                    filepath = str(DOCS_DIR / d["fichier"])
            if os.path.exists(filepath):
                os.unlink(filepath)
            data["documents"] = [x for x in data["documents"] if x["id"] != doc_id]
            save_db(data)
            return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Document non trouve")

# ============ EMAILING ENDPOINTS ============
@app.post("/api/email/send")
async def send_email(req: EmailSend, user=Depends(require_referent)):
    data = load_db()
    recipients = list(req.recipients)
    
    # Collect emails from contact_ids
    for cid in req.contact_ids:
        for c in data["contacts"]:
            if c["id"] == cid and c.get("email"):
                recipients.append(c["email"])
    
    recipients = list(set(r for r in recipients if r))
    if not recipients:
        raise HTTPException(status_code=400, detail="Aucun destinataire valide")
    
    try:
        send_email_smtp(recipients, req.subject, req.body, html=req.html)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur envoi email: {str(e)}")
    
    # Log email in each contact's history
    for cid in req.contact_ids:
        for c in data["contacts"]:
            if c["id"] == cid:
                add_history(c, user, "email", f"Sujet: {req.subject} | A: {', '.join(recipients)}")
    save_db(data)
    
    return {"status": "ok", "sent_to": recipients, "count": len(recipients)}

@app.get("/api/email/templates")
async def get_templates(user=Depends(require_referent)):
    return {
        "newsletter": {
            "subject": "Newsletter - Les Ami(e)s de Romy",
            "body": "Bonjour,\n\nVoici les dernieres nouvelles de l'association Les Ami(e)s de Romy.\n\n[Contenu a personnaliser]\n\nCordialement,\nL'equipe des Ami(e)s de Romy"
        },
        "relance": {
            "subject": "Relance - Les Ami(e)s de Romy",
            "body": "Bonjour,\n\nNous vous contactons dans le cadre de l'association Les Ami(e)s de Romy.\n\n[Contenu a personnaliser]\n\nCordialement,\nL'equipe des Ami(e)s de Romy"
        },
        "invitation": {
            "subject": "Invitation evenement - Les Ami(e)s de Romy",
            "body": "Bonjour,\n\nVous etes invite(e) a notre prochain evenement.\n\nDate: [A completer]\nLieu: [A completer]\n\n[Description]\n\nAu plaisir de vous y retrouver,\nL'equipe des Ami(e)s de Romy"
        },
        "newsletter_html": {
            "subject": "Newsletter - Les Ami(e)s de Romy",
            "body": "<div style=\"font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#f3e0f5;padding:20px;border-radius:12px\"><div style=\"text-align:center;padding:20px 0\"><span style=\"font-size:2rem\">&hearts;</span><h1 style=\"color:#d03ec6;margin:0\">Les Ami(e)s de Romy</h1></div><div style=\"background:#fff;padding:20px;border-radius:8px\"><p>Bonjour,</p><p>Voici les dernieres nouvelles de l association.</p><p>[Contenu a personnaliser]</p><p style=\"color:#999;font-size:.85rem\">L equipe des Ami(e)s de Romy</p></div></div>",
            "html": True
        },
        "remerciement_don": {
            "subject": "Merci pour votre don - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nNous tenons a vous remercier chaleureusement pour votre don de [MONTANT] euros en faveur de notre association Les Ami(e)s de Romy.\n\nGrace a votre generosite, nous pouvons poursuivre nos actions et accompagner nos projets tout au long de l annee [ANNEE]. Votre soutien est essentiel pour nous.\n\nUn recu fiscal vous sera adresse prochainement si votre don ouvre droit a une deduction fiscale.\n\nEncore merci pour votre confiance et votre fidelite.\n\nCordialement,\nL equipe des Ami(e)s de Romy",
            "html": False
        },
        "convocation_ag": {
            "subject": "Convocation a l'Assemblee Generale - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nL equipe de l association Les Ami(e)s de Romy a le plaisir de vous convoquer a l Assemblee Generale de l annee [ANNEE].\n\nDate: [A completer]\nLieu: [A completer]\nOrdre du jour:\n  - Rapport moral du President\n  - Rapport financier du Tresorier\n  - Renouvellement du Conseil d Administration\n  - Questions diverses\n\nVotre presence est importante pour la vie de notre association. En cas d empechement, vous pouvez vous faire representer par le biais d un pouvoir.\n\nMerci de confirmer votre presence aupres du secretaire.\n\nCordialement,\nLe Conseil d Administration des Ami(e)s de Romy",
            "html": False
        },
        "bilan_annuel": {
            "subject": "Bilan annuel [ANNEE] - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nNous avons le plaisir de vous presenter le bilan annuel de l association Les Ami(e)s de Romy pour l annee [ANNEE].\n\nCette annee a ete riche en evenements et en actions:\n  - Nombre d adherents: [A completer]\n  - Evenements organises: [A completer]\n  - Montant des dons recoltes: [MONTANT] euros\n  - Actions mennees: [A completer]\n\nNous tenons a remercier l ensemble de nos benevoles, nos donateurs et nos adherents pour leur soutien precieux.\n\nVous trouverez en piece jointe le rapport complet de nos activites.\n\nCordialement,\nL equipe des Ami(e)s de Romy",
            "html": False
        },
        "relance_cotisation": {
            "subject": "Relance de cotisation [ANNEE] - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nNous vous contactons concernant votre cotisation annuelle a l association Les Ami(e)s de Romy pour l annee [ANNEE].\n\nA ce jour, nous n avons pas recu le reglement de votre cotisation d un montant de [MONTANT] euros. Nous vous invitons a regulariser votre situation dans les meilleurs delais.\n\nPour cela, vous pouvez effectuer votre paiement:\n  - Par cheque a l ordre de Les Ami(e)s de Romy\n  - Par virement bancaire (contactez-nous pour le RIB)\n  - En especes lors d un deplacement a l association\n\nVotre cotisation est essentielle au bon fonctionnement de nos actions et nous comptons sur votre soutien.\n\nSi vous avez deja regle votre cotisation, merci d ignorer ce message et d en excuser la relance.\n\nCordialement,\nLe Tresorier des Ami(e)s de Romy",
            "html": False
        },
        "anniversaire_adhesion": {
            "subject": "Felicitations pour votre anniversaire d'adhesion - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nL equipe des Ami(e)s de Romy est heureuse de vous feliciter a l occasion de l anniversaire de votre adhesion a notre association.\n\nDepuis votre adhesion, vous contribuez a la dynamique et au rayonnement de notre association grace a votre engagement et votre fidelite.\n\nNous esperons continuer a vous compter parmi nos adherents pour les annees a venir et a partager ensemble de nouveaux projets enrichissants.\n\nToute l equipe vous souhaite une excellente continuation et vous remercie pour votre soutien continu.\n\nCordialement,\nL equipe des Ami(e)s de Romy",
            "html": False
        },
        "merci_benevole": {
            "subject": "Remerciement pour votre benevolat - Les Ami(e)s de Romy",
            "body": "Bonjour [PRENOM] [NOM],\n\nL equipe des Ami(e)s de Romy tient a vous adreser ses plus vifs remerciements pour votre engagement benevole au sein de notre association.\n\nVotre devouement et votre disponibilite tout au long de l annee [ANNEE] ont largement contribue au succes de nos actions et evenements. Grace a vous, nous avons pu atteindre nos objectifs et offrir un accompagnement de qualite.\n\nNous tenons a souligner votre implication notamment sur les projets suivants:\n  - [A completer]\n  - [A completer]\n\nNous esperons pouvoir compter de nouveau sur votre soutien pour les prochaines actions. N hesitez pas a nous faire part de vos suggestions ou envies pour les projets a venir.\n\nEncore merci pour tout ce que vous faites.\n\nCordialement,\nL equipe des Ami(e)s de Romy",
            "html": False
        }
    }

@app.get("/api/email/preview-contacts")
async def preview_email_contacts(contact_ids: str = "", user=Depends(require_referent)):
    """Returns emails for given contact IDs"""
    data = load_db()
    ids = [x.strip() for x in contact_ids.split(",") if x.strip()]
    emails = []
    for c in data["contacts"]:
        if c["id"] in ids and c.get("email"):
            emails.append({"id": c["id"], "nom": f"{c['prenom']} {c['nom']}", "email": c["email"]})
    return emails

# ============ STATS ============
@app.get("/api/stats")
async def get_stats(user=Depends(require_referent)):
    data = load_db()
    contacts = data["contacts"]
    qualites = {}
    for c in contacts:
        q = c.get("qualite", "non specifie") or "non specifie"
        qualites[q] = qualites.get(q, 0) + 1
    return {
        "total_contacts": len(contacts),
        "total_evenements": len(data.get("evenements", [])),
        "total_documents": len(data.get("documents", [])),
        "total_cotisations": len(data.get("cotisations", [])),
        "par_qualite": qualites,
        "avec_email": sum(1 for c in contacts if c.get("email")),
        "sans_email": sum(1 for c in contacts if not c.get("email")),
        "avec_telephone": sum(1 for c in contacts if c.get("telephone")),
        "cotisations_payees": sum(1 for c in data.get("cotisations", []) if c.get("statut") == "paye"),
        "cotisations_impayees": sum(1 for c in data.get("cotisations", []) if c.get("statut") != "paye"),
        "montant_cotisations": sum(c.get("montant", 0) for c in data.get("cotisations", []) if c.get("statut") == "paye"),
        "total_dons": len(data.get("dons", [])),
        "montant_dons": sum(d.get("montant", 0) for d in data.get("dons", [])),
        "total_taches": len(data.get("taches", [])),
        "taches_a_faire": sum(1 for t in data.get("taches", []) if t.get("statut") == "a_faire"),
        "taches_terminees": sum(1 for t in data.get("taches", []) if t.get("statut") == "terminee"),
    }

# ============ STATS AVANCEES ============
@app.get("/api/stats/advanced")
async def get_stats_advanced(user=Depends(require_referent)):
    data = load_db()
    contacts = data["contacts"]
    events = data.get("evenements", [])
    cots = data.get("cotisations", [])

    # Cotisations par annee
    cotisations_par_annee = {}
    for c in cots:
        annee = c.get("annee", "N/A")
        if annee not in cotisations_par_annee:
            cotisations_par_annee[annee] = {"total": 0, "paye": 0, "impaye": 0, "montant": 0}
        cotisations_par_annee[annee]["total"] += 1
        if c.get("statut") == "paye":
            cotisations_par_annee[annee]["paye"] += 1
            cotisations_par_annee[annee]["montant"] += c.get("montant", 0)
        else:
            cotisations_par_annee[annee]["impaye"] += 1

    # Participation aux evenements
    participation = []
    contact_by_id = {c["id"]: c for c in contacts}
    for ev in events:
        parts = ev.get("participants", [])
        presence = ev.get("presence", {})
        presents = sum(1 for v in presence.values() if v == "present")
        absents = sum(1 for v in presence.values() if v == "absent")
        excuses = sum(1 for v in presence.values() if v == "excuse")
        participation.append({
            "titre": ev.get("titre", ""),
            "date": ev.get("date", ""),
            "inscrits": len(parts),
            "presents": presents,
            "absents": absents,
            "excuses": excuses,
            "taux_presence": round(presents / len(parts) * 100, 1) if parts else 0
        })

    # Repartition par tags
    tags_count = {}
    for c in contacts:
        for tag in c.get("tags", "").split(","):
            tag = tag.strip()
            if tag:
                tags_count[tag] = tags_count.get(tag, 0) + 1

    # Contacts sans cotisation (annee en cours)
    current_year = str(datetime.now().year)
    contacts_avec_cot_annee = set()
    for c in cots:
        if c.get("annee") == current_year and c.get("statut") == "paye":
            contacts_avec_cot_annee.add(c.get("contact_id"))
    contacts_sans_cot = [c["id"] for c in contacts if c["id"] not in contacts_avec_cot_annee]

    return {
        "cotisations_par_annee": cotisations_par_annee,
        "participation_evenements": participation,
        "tags_count": tags_count,
        "contacts_sans_cotisation": len(contacts_sans_cot),
        "total_cotisations": len(cots),
        "total_evenements": len(events),
    }

# ============ CALENDRIER MENSUEL ============
@app.get("/api/evenements/calendrier")
async def get_calendrier(mois: int = 0, user=Depends(require_auth)):
    """Retourne les evenements du mois specifie (0=mois courant, 1=mois prochain, -1=mois precedent)."""
    from datetime import timedelta
    data = load_db()
    today = datetime.now().date()
    # Calculer le mois cible
    if mois == 0:
        target = today.replace(day=1)
    elif mois > 0:
        target = today.replace(day=1)
        for _ in range(mois):
            if target.month == 12:
                target = target.replace(year=target.year + 1, month=1)
            else:
                target = target.replace(month=target.month + 1)
    else:
        target = today.replace(day=1)
        for _ in range(abs(mois)):
            if target.month == 1:
                target = target.replace(year=target.year - 1, month=12)
            else:
                target = target.replace(month=target.month - 1)

    # Filtrer les evenements du mois
    events_mois = []
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for ev in data.get("evenements", []):
        try:
            ev_date = datetime.strptime(ev.get("date", ""), "%Y-%m-%d").date()
        except Exception:
            continue
        if ev_date.year == target.year and ev_date.month == target.month:
            # Resoudre les noms des participants
            parts_noms = []
            for pid in ev.get("participants", []):
                c = contact_by_id.get(pid, {})
                parts_noms.append(f"{c.get('prenom', '')} {c.get('nom', '')}".strip())
            events_mois.append({
                "id": ev["id"],
                "titre": ev.get("titre", ""),
                "date": ev.get("date", ""),
                "heure": ev.get("heure", ""),
                "lieu": ev.get("lieu", ""),
                "description": ev.get("description", ""),
                "participants": parts_noms,
                "jour": ev_date.day
            })

    # Informations sur le mois
    if target.month == 12:
        next_month = target.replace(year=target.year + 1, month=1)
    else:
        next_month = target.replace(month=target.month + 1)
    prev_month = target.replace(day=1) - timedelta(days=1)
    prev_month = prev_month.replace(day=1)

    return {
        "mois_cible": target.strftime("%Y-%m"),
        "mois_label": target.strftime("%B %Y"),
        "mois_label_fr": _format_date_fr(target.strftime("%Y-%m-01")).split()[1] + " " + str(target.year),
        "nb_jours": (next_month - target).days,
        "events": events_mois,
        "prev_mois": prev_month.strftime("%Y-%m"),
        "next_mois": next_month.strftime("%Y-%m"),
    }

# ============ TACHES ENDPOINTS ============
@app.get("/api/taches")
async def list_taches(statut: str = "", user=Depends(require_referent)):
    data = load_db()
    taches = data.get("taches", [])
    if statut:
        taches = [t for t in taches if t.get("statut") == statut]
    return taches

@app.post("/api/taches")
async def create_tache(req: TacheCreate, user=Depends(require_referent)):
    data = load_db()
    tache = {
        "id": f"tache-{secrets.token_hex(8)}",
        "titre": req.titre,
        "description": req.description,
        "assigne_a": req.assigne_a,
        "echeance": req.echeance,
        "priorite": req.priorite,
        "statut": "a_faire",
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data.setdefault("taches", []).append(tache)
    save_db(data)
    backup_db()
    return {"status": "ok", "tache": tache}

@app.put("/api/taches/{tache_id}")
async def update_tache(tache_id: str, req: TacheUpdate, user=Depends(require_referent)):
    data = load_db()
    for t in data.get("taches", []):
        if t["id"] == tache_id:
            if req.titre is not None: t["titre"] = req.titre
            if req.description is not None: t["description"] = req.description
            if req.assigne_a is not None: t["assigne_a"] = req.assigne_a
            if req.echeance is not None: t["echeance"] = req.echeance
            if req.priorite is not None: t["priorite"] = req.priorite
            if req.statut is not None: t["statut"] = req.statut
            save_db(data)
            backup_db()
            return {"status": "ok", "tache": t}
    raise HTTPException(status_code=404, detail="Tache non trouvee")

@app.delete("/api/taches/{tache_id}")
async def delete_tache(tache_id: str, user=Depends(require_referent)):
    data = load_db()
    data["taches"] = [t for t in data.get("taches", []) if t["id"] != tache_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}

# ============ DONS / SPONSORS ENDPOINTS ============
@app.get("/api/dons")
async def list_dons(type_don: str = "", user=Depends(require_referent)):
    data = load_db()
    dons = data.get("dons", [])
    if type_don:
        dons = [d for d in dons if d.get("type_don") == type_don]
    return dons

@app.post("/api/dons")
async def create_don(req: DonCreate, user=Depends(require_referent)):
    data = load_db()
    don = {
        "id": f"don-{secrets.token_hex(8)}",
        "contact_id": req.contact_id,
        "nom": req.nom,
        "montant": req.montant,
        "date": req.date,
        "mode_paiement": req.mode_paiement,
        "type_don": req.type_don,
        "notes": req.notes,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data.setdefault("dons", []).append(don)
    save_db(data)
    backup_db()
    return {"status": "ok", "don": don}

@app.delete("/api/dons/{don_id}")
async def delete_don(don_id: str, user=Depends(require_referent)):
    data = load_db()
    data["dons"] = [d for d in data.get("dons", []) if d["id"] != don_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}

# ============ PHOTOS EVENEMENTS ============
@app.post("/api/evenements/{evt_id}/photos")
async def upload_event_photo(evt_id: str, request: Request, user=Depends(require_referent)):
    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier requis")
    ALLOWED_IMG = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_IMG:
        raise HTTPException(status_code=400, detail=f"Type non autorise: {ext}")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Max 10 Mo")
    safe_name = re.sub(r'[^\w\.-]', '_', file.filename)
    safe_name = safe_name.replace("..", "").replace("/", "").replace("\\", "")
    # Prefix avec evt_id pour organiser
    photo_filename = f"{evt_id}_{safe_name}"
    filepath = DOCS_DIR / "event_photos" / photo_filename
    filepath.parent.mkdir(exist_ok=True)
    with open(filepath, "wb") as f:
        f.write(content)
    data = load_db()
    for ev in data.get("evenements", []):
        if ev["id"] == evt_id:
            ev.setdefault("photos", []).append({
                "filename": photo_filename,
                "filepath": str(filepath),
                "cree_par": user["username"],
                "cree_le": now_iso()
            })
            save_db(data)
            backup_db()
            return {"status": "ok", "photo": photo_filename}
    raise HTTPException(status_code=404, detail="Evenement non trouve")

@app.get("/api/evenements/{evt_id}/photos")
async def list_event_photos(evt_id: str, user=Depends(require_auth)):
    data = load_db()
    for ev in data.get("evenements", []):
        if ev["id"] == evt_id:
            return ev.get("photos", [])
    raise HTTPException(status_code=404, detail="Evenement non trouve")

@app.delete("/api/evenements/{evt_id}/photos/{photo_idx}")
async def delete_event_photo(evt_id: str, photo_idx: int, user=Depends(require_referent)):
    data = load_db()
    for ev in data.get("evenements", []):
        if ev["id"] == evt_id:
            photos = ev.get("photos", [])
            if 0 <= photo_idx < len(photos):
                if os.path.exists(photos[photo_idx].get("filepath", "")):
                    os.unlink(photos[photo_idx]["filepath"])
                photos.pop(photo_idx)
                save_db(data)
                backup_db()
                return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Photo non trouvee")

# ============ EXPORT EXCEL ============
@app.get("/api/export/contacts-excel")
async def export_contacts_excel(user=Depends(require_referent)):
    import openpyxl
    import tempfile
    data = load_db()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    headers = ["ID", "Prenom", "Nom", "Qualite", "Telephone", "Email", "Tags", "Notes"]
    ws.append(headers)
    for c in data["contacts"]:
        ws.append([c.get("id",""), c.get("prenom",""), c.get("nom",""), c.get("qualite",""),
                    c.get("telephone",""), c.get("email",""), c.get("tags",""), c.get("notes","")])
    # Cotisations sheet
    ws2 = wb.create_sheet("Cotisations")
    ws2.append(["Contact", "Annee", "Montant", "Statut", "Mode", "Date paiement"])
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for cot in data.get("cotisations", []):
        c = contact_by_id.get(cot.get("contact_id"), {})
        ws2.append([f"{c.get('prenom','')} {c.get('nom','')}", cot.get("annee",""), cot.get("montant",0),
                    cot.get("statut",""), cot.get("mode_paiement",""), cot.get("date_paiement","")])
    # Evenements sheet
    ws3 = wb.create_sheet("Evenements")
    ws3.append(["Titre", "Date", "Heure", "Lieu", "Description", "Participants", "Cree par"])
    for ev in data.get("evenements", []):
        parts = []
        for pid in ev.get("participants", []):
            c = contact_by_id.get(pid, {})
            parts.append(f"{c.get('prenom','')} {c.get('nom','')}".strip())
        ws3.append([ev.get("titre",""), ev.get("date",""), ev.get("heure",""),
                    ev.get("lieu",""), ev.get("description",""), ", ".join(parts), ev.get("cree_par","")])
    tmp_path = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp_path)
    return FileResponse(tmp_path, filename="export_contacts_cotisations.xlsx")

# ============ EXPORT ICS ============
@app.get("/api/export/evenements-ics")
async def export_evenements_ics(user=Depends(require_auth)):
    import tempfile
    data = load_db()
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//CRM Romy//FR", "CALSCALE:GREGORIAN"]
    for ev in data.get("evenements", []):
        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{ev['id']}@crm-romy")
        date = ev.get("date", "")
        heure = ev.get("heure", "")
        if date:
            if heure:
                dt_start = f"{date.replace('-','')}T{heure.replace(':','')}00"
                lines.append(f"DTSTART:{dt_start}")
                lines.append(f"DTEND:{date.replace('-','')}T{heure.replace(':','')}00")
            else:
                lines.append(f"DTSTART;VALUE=DATE:{date.replace('-','')}")
        lines.append(f"SUMMARY:{ev.get('titre','')}")
        if ev.get("lieu"):
            lines.append(f"LOCATION:{ev['lieu']}")
        if ev.get("description"):
            lines.append(f"DESCRIPTION:{ev['description']}")
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    ics_content = (chr(13) + chr(10)).join(lines)
    tmp_path = tempfile.mktemp(suffix=".ics")
    with open(tmp_path, "w", encoding="utf-8") as f:
        f.write(ics_content)
    return FileResponse(tmp_path, filename="evenements_romy.ics", media_type="text/calendar")

# ============ FEEDBACK (BUG / SUGGESTION) ============
@app.post("/api/feedback")
async def create_feedback(req: FeedbackCreate, user=Depends(require_auth)):
    data = load_db()
    fb = {
        "id": f"fb-{secrets.token_hex(8)}",
        "type": req.type,
        "message": req.message,
        "page": req.page,
        "user": user["username"],
        "date": now_iso(),
        "statut": "ouvert"
    }
    data.setdefault("feedbacks", []).append(fb)
    save_db(data)
    backup_db()
    # Envoyer un email a Vincent
    smtp = load_smtp()
    if smtp.get("password"):
        sujet = f"[CRM Romy] {req.type.upper()} - par {user['username']}"
        body = f"Type: {req.type}\nUtilisateur: {user['username']}\nPage: {req.page or 'non precisee'}\n\nMessage:\n{req.message}"
        try:
            send_email_smtp(["vincentminvielle@gmail.com"], sujet, body)
        except Exception:
            pass  # Ne pas bloquer si l email echoue
    return {"status": "ok", "feedback": fb}

@app.get("/api/feedback")
async def list_feedback(user=Depends(require_referent)):
    data = load_db()
    return data.get("feedbacks", [])

@app.put("/api/feedback/{fb_id}")
async def update_feedback(fb_id: str, req: dict, user=Depends(require_referent)):
    data = load_db()
    for fb in data.get("feedbacks", []):
        if fb["id"] == fb_id:
            if "statut" in req:
                fb["statut"] = req["statut"]
            save_db(data)
            backup_db()
            return {"status": "ok", "feedback": fb}
    raise HTTPException(status_code=404, detail="Feedback non trouve")

@app.delete("/api/feedback/{fb_id}")
async def delete_feedback(fb_id: str, user=Depends(require_referent)):
    data = load_db()
    data["feedbacks"] = [f for f in data.get("feedbacks", []) if f["id"] != fb_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}

# ============ CARTE DE VISITE PUBLIQUE ============
@app.get("/api/public/carte", response_class=HTMLResponse)
async def carte_visite_publique():
    """Page de carte de visite publique de l'association — aucune auth requise."""
    data = load_db()
    evenements = data.get("evenements", [])
    # Trouver le prochain événement (date >= aujourd'hui)
    prochain_evt = None
    today_str = datetime.now().strftime("%Y-%m-%d")
    for ev in evenements:
        ev_date = ev.get("date", "")
        if ev_date and ev_date >= today_str:
            if prochain_evt is None or ev_date < prochain_evt.get("date", "9999"):
                prochain_evt = ev
    # Infos association (depuis db.json si disponible, sinon valeurs par défaut)
    asso_info = data.get("association", {})
    nom_asso = "Les Ami(e)s de Romy"
    slogan = asso_info.get("slogan", "Sensibiliser, protéger et accompagner")
    mission = asso_info.get("mission", "Notre association accompagne les familles touchées par la maladie de Huntington en proposant soutien, activités et répit. Nous sensibilisons le grand public et defendons les droits des malades et de leurs proches.")
    email = asso_info.get("email", "contact@amisderomy.fr")
    tel = asso_info.get("tel", "06 00 00 00 00")
    adresse = asso_info.get("adresse", "France")

    helloasso_url = "https://www.helloasso.com/associations/les-ami-es-de-romy"

    # Section prochain événement
    if prochain_evt:
        evt_html = f"""
        <div class="carte-evt">
          <div class="evt-badge">Prochain événement</div>
          <div class="evt-titre">{prochain_evt.get('titre', '')}</div>
          <div class="evt-date">📅 {prochain_evt.get('date', '')}{(' à ' + prochain_evt.get('heure', '')) if prochain_evt.get('heure') else ''}</div>
          {f'<div class="evt-lieu">📍 {prochain_evt.get("lieu", "")}</div>' if prochain_evt.get('lieu') else ''}
          {f'<div class="evt-desc">{prochain_evt.get("description", "")}</div>' if prochain_evt.get('description') else ''}
        </div>
        """
    else:
        evt_html = '<div class="carte-evt"><div class="evt-badge">Prochain événement</div><div class="evt-none">Aucun événement planifié pour le moment — revenez bientôt !</div></div>'

    html_page = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Les Ami(e)s de Romy — Carte de visite</title>
<meta property="og:title" content="Les Ami(e)s de Romy">
<meta property="og:description" content="{slogan}">
<meta property="og:image" content="/static/romy_banner_new.png">
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
:root {{
  --bleu-marine: #0a335c;
  --vert-sauge: #6c9186;
  --rose-pale: #dd83a9;
  --bg: #f4f6f8;
  --surface: #ffffff;
  --text: #2a2a2a;
  --text-light: #666;
  --radius: 16px;
}}
body {{
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  background: linear-gradient(135deg, var(--bleu-marine) 0%, var(--vert-sauge) 100%);
  min-height: 100vh;
  display: flex;
  align-items: center;
  justify-content: center;
  padding: 1rem;
}}
.carte {{
  background: var(--surface);
  border-radius: var(--radius);
  max-width: 480px;
  width: 100%;
  overflow: hidden;
  box-shadow: 0 20px 60px rgba(0,0,0,0.3);
  animation: fadeIn 0.6s ease;
}}
@keyframes fadeIn {{
  from {{ opacity: 0; transform: translateY(20px); }}
  to {{ opacity: 1; transform: translateY(0); }}
}}
.carte-header {{
  background: linear-gradient(135deg, var(--bleu-marine), var(--vert-sauge));
  padding: 2rem 1.5rem 1.5rem;
  text-align: center;
  color: #fff;
}}
.carte-logo {{
  width: 120px;
  height: 120px;
  border-radius: 50%;
  border: 4px solid rgba(255,255,255,0.3);
  object-fit: cover;
  margin: 0 auto .8rem;
  display: block;
  background: #fff;
}}
.carte-nom {{
  font-size: 1.5rem;
  font-weight: 700;
  margin-bottom: .3rem;
}}
.carte-slogan {{
  font-size: .95rem;
  opacity: 0.9;
  font-style: italic;
}}
.carte-body {{ padding: 1.5rem; }}
.carte-mission {{
  font-size: .9rem;
  color: var(--text);
  line-height: 1.6;
  margin-bottom: 1.5rem;
  padding: 1rem;
  background: var(--bg);
  border-radius: 12px;
  border-left: 4px solid var(--vert-sauge);
}}
.carte-coords {{
  display: flex;
  flex-direction: column;
  gap: .6rem;
  margin-bottom: 1.5rem;
}}
.coord {{
  display: flex;
  align-items: center;
  gap: .6rem;
  font-size: .9rem;
  color: var(--text);
}}
.coord-icon {{
  width: 36px; height: 36px;
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  background: var(--bg);
  font-size: 1rem;
  flex-shrink: 0;
}}
.coord a {{ color: var(--bleu-marine); text-decoration: none; }}
.coord a:hover {{ text-decoration: underline; }}
.carte-evt {{
  margin-bottom: 1.5rem;
  padding: 1rem;
  background: linear-gradient(135deg, rgba(10,51,92,0.05), rgba(108,145,134,0.05));
  border-radius: 12px;
  border: 1px solid rgba(10,51,92,0.1);
}}
.evt-badge {{
  display: inline-block;
  background: var(--vert-sauge);
  color: #fff;
  font-size: .7rem;
  font-weight: 700;
  padding: .2rem .6rem;
  border-radius: 20px;
  margin-bottom: .5rem;
  text-transform: uppercase;
  letter-spacing: .5px;
}}
.evt-titre {{ font-weight: 700; color: var(--bleu-marine); margin-bottom: .3rem; font-size: 1.05rem; }}
.evt-date {{ font-size: .85rem; color: var(--text); margin-bottom: .2rem; }}
.evt-lieu {{ font-size: .85rem; color: var(--text-light); margin-bottom: .2rem; }}
.evt-desc {{ font-size: .8rem; color: var(--text-light); line-height: 1.5; }}
.evt-none {{ font-size: .85rem; color: var(--text-light); font-style: italic; }}
.carte-qr {{
  text-align: center;
  margin-bottom: 1.5rem;
}}
.carte-qr-label {{
  font-size: .8rem;
  color: var(--text-light);
  margin-bottom: .5rem;
  font-weight: 600;
}}
.qr-box {{
  display: inline-block;
  padding: .6rem;
  background: var(--surface);
  border: 2px solid var(--bleu-marine);
  border-radius: 12px;
}}
.qr-box img {{ width: 140px; height: 140px; display: block; }}
.carte-social {{
  display: flex;
  justify-content: center;
  gap: .6rem;
  flex-wrap: wrap;
  margin-bottom: 1rem;
}}
.btn-social {{
  display: inline-flex;
  align-items: center;
  gap: .4rem;
  padding: .6rem 1.2rem;
  border-radius: 30px;
  font-size: .85rem;
  font-weight: 600;
  text-decoration: none;
  transition: transform .2s, box-shadow .2s;
}}
.btn-social:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.15); }}
.btn-helloasso {{ background: var(--bleu-marine); color: #fff; }}
.btn-mail {{ background: var(--vert-sauge); color: #fff; }}
.btn-share {{ background: var(--rose-pale); color: #fff; }}
.carte-footer {{
  text-align: center;
  padding: 1rem;
  font-size: .75rem;
  color: var(--text-light);
  border-top: 1px solid #eee;
}}
.carte-footer a {{ color: var(--bleu-marine); }}
@media (max-width: 520px) {{
  .carte-logo {{ width: 90px; height: 90px; }}
  .carte-nom {{ font-size: 1.25rem; }}
  .qr-box img {{ width: 120px; height: 120px; }}
  .btn-social {{ padding: .5rem .9rem; font-size: .8rem; }}
}}
</style>
</head>
<body>
<div class="carte">
  <div class="carte-header">
    <img class="carte-logo" src="/static/romy_banner_new.png" alt="Logo Les Ami(e)s de Romy">
    <div class="carte-nom">{nom_asso}</div>
    <div class="carte-slogan">« {slogan} »</div>
  </div>
  <div class="carte-body">
    <div class="carte-mission">{mission}</div>
    <div class="carte-coords">
      <div class="coord"><div class="coord-icon">✉️</div><a href="mailto:{email}">{email}</a></div>
      <div class="coord"><div class="coord-icon">📞</div><a href="tel:{tel.replace(' ', '')}">{tel}</a></div>
      <div class="coord"><div class="coord-icon">📍</div><span>{adresse}</span></div>
    </div>
    {evt_html}
    <div class="carte-qr">
      <div class="carte-qr-label">Scannez pour soutenir l'association</div>
      <div class="qr-box">
        <img src="https://api.qrserver.com/v1/create-qr-code/?size=140x140&data={helloasso_url}" alt="QR Code HelloAsso">
      </div>
    </div>
    <div class="carte-social">
      <a class="btn-social btn-helloasso" href="{helloasso_url}" target="_blank">❤️ Soutenir sur HelloAsso</a>
      <a class="btn-social btn-mail" href="mailto:{email}">✉️ Nous écrire</a>
      <a class="btn-social btn-share" href="javascript:void(0)" onclick="navigator.share?navigator.share({{title:'Les Ami(e)s de Romy',url:window.location.href}}):window.open(window.location.href)">📤 Partager</a>
    </div>
  </div>
  <div class="carte-footer">
    © {datetime.now().strftime('%Y')} Les Ami(e)s de Romy — CRM v{VERSION}
  </div>
</div>
</body>
</html>"""
    return HTMLResponse(html_page)


# ============ RECHERCHE GLOBALE ============
@app.get("/api/search")
async def global_search(q: str = Query("", min_length=0), user=Depends(require_auth)):
    """Recherche globale sur tous les types de donnees.
    Retourne une liste unifiee de resultats (type, id, titre, sous_titre, view, detail_fn).
    Limite a 20 resultats. Les membres ne voient que les contacts autorises.
    """
    ql = q.strip().lower()
    if not ql:
        return []
    data = load_db()
    results = []
    member = is_member(user)

    # --- Contacts: nom/prenom/email/telephone ---
    for c in data.get("contacts", []):
        if member and c.get("qualite", "").lower() not in ("membre asso", "referent", "benevole"):
            continue
        hay = (c.get("prenom", "") + " " + c.get("nom", "") + " " + c.get("email", "") + " " + c.get("telephone", "")).lower()
        if ql in hay:
            results.append({
                "type": "contact",
                "id": c["id"],
                "titre": f"{c.get('prenom', '')} {c.get('nom', '')}".strip(),
                "sous_titre": c.get("qualite", "") or c.get("email", "") or c.get("telephone", "") or "-",
                "view": "contacts",
                "icon": "\U0001F464"
            })

    # --- Evenements: titre/lieu ---
    for ev in data.get("evenements", []):
        hay = (ev.get("titre", "") + " " + ev.get("lieu", "")).lower()
        if ql in hay:
            results.append({
                "type": "evenement",
                "id": ev["id"],
                "titre": ev.get("titre", "") or "Evenement",
                "sous_titre": f"{ev.get('date', '-')} {ev.get('lieu', '')}".strip(),
                "view": "evenements",
                "icon": "\U0001F4C5"
            })

    # --- Notes de frais: description ---
    for n in data.get("notes_frais", []):
        if member:
            continue  # notes-frais reserved to referents
        hay = (n.get("description", "") + " " + n.get("categorie", "")).lower()
        if ql in hay:
            results.append({
                "type": "note_frais",
                "id": n["id"],
                "titre": n.get("description", "") or "Note de frais",
                "sous_titre": f"{n.get('date', '-')} | {n.get('montant', '-')} EUR | {n.get('contact_nom', '-')}",
                "view": "frais",
                "icon": "\U0001F4B5"
            })

    # --- Votes: titre ---
    for v in data.get("votes", []):
        if member:
            continue
        hay = v.get("titre", "").lower()
        if ql in hay:
            results.append({
                "type": "vote",
                "id": v["id"],
                "titre": v.get("titre", "") or "Vote",
                "sous_titre": f"{v.get('date', '-')} | {v.get('type_vote', '-')}",
                "view": "votes",
                "icon": "\U0001F3F3"
            })

    # --- Comptes-rendus: ordre_du_jour + titre ---
    for cr in data.get("comptes_rendus", []):
        hay = (cr.get("titre", "") + " " + cr.get("ordre_du_jour", "")).lower()
        if ql in hay:
            results.append({
                "type": "cr",
                "id": cr["id"],
                "titre": cr.get("titre", "") or "Compte-rendu",
                "sous_titre": f"{cr.get('date', '-')} | {cr.get('type_reunion', '-')}",
                "view": "crs",
                "icon": "\U0001F4DD"
            })

    # --- Accompagnements: nom du contact (resolu) ---
    if not member:
        contact_by_id = {c["id"]: c for c in data.get("contacts", [])}
        for a in data.get("accompagnements", []):
            c = contact_by_id.get(a.get("contact_id", ""))
            nom = f"{c.get('prenom', '')} {c.get('nom', '')}".strip() if c else ""
            hay = (nom + " " + a.get("type_suivi", "")).lower()
            if ql in hay:
                results.append({
                    "type": "accompagnement",
                    "id": a["id"],
                    "titre": nom or a.get("type_suivi", "") or "Accompagnement",
                    "sous_titre": f"{a.get('statut', '-')} | {a.get('type_suivi', '-')} | {a.get('date_debut', '-')}",
                    "view": "accompagnements",
                    "icon": "\U0001F46A"
                })

    # --- Documents: nom ---
    if not member:
        for d in data.get("documents", []):
            hay = (d.get("nom", "") + " " + d.get("type_doc", "") + " " + d.get("description", "")).lower()
            if ql in hay:
                results.append({
                    "type": "document",
                    "id": d["id"],
                    "titre": d.get("nom", "") or "Document",
                    "sous_titre": f"{d.get('type_doc', '-')} | {d.get('cree_par', '-')}",
                    "view": "documents",
                    "icon": "\U0001F4C4"
                })

    # --- Dons: donateur (nom) ---
    if not member:
        for d in data.get("dons", []):
            hay = (d.get("nom", "") + " " + d.get("type_don", "") + " " + d.get("notes", "")).lower()
            if ql in hay:
                results.append({
                    "type": "don",
                    "id": d["id"],
                    "titre": d.get("nom", "") or "Don",
                    "sous_titre": f"{d.get('montant', '-')} EUR | {d.get('date', '-')} | {d.get('type_don', '-')}",
                    "view": "dons",
                    "icon": "\U0001F381"
                })

    # --- Taches: titre ---
    if not member:
        for t in data.get("taches", []):
            hay = (t.get("titre", "") + " " + t.get("description", "")).lower()
            if ql in hay:
                results.append({
                    "type": "tache",
                    "id": t["id"],
                    "titre": t.get("titre", "") or "Tache",
                    "sous_titre": f"{t.get('priorite', '-')} | {t.get('statut', '-')}",
                    "view": "taches",
                    "icon": "\u2705"
                })

    # Limite globale a 20 resultats
    return results[:20]


# ============ STATIC FILES + SPA ============
@app.get("/")
async def index():
    index_path = BASE_DIR / "static" / "index.html"
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

@app.get("/favicon.ico")
async def favicon():
    return Response(b"", media_type="image/x-icon")

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
# Monter uploads en statique pour accès aux justificatifs et photos événements
try:
    app.mount("/uploads", StaticFiles(directory=str(DOCS_DIR)), name="uploads")
except RuntimeError:
    pass  # déjà monté ou dossier manquant

# Endpoint pour visualiser une facture depuis le stockage Nipogi
@app.get("/api/frais/image/{filename}")
async def get_facture_image(filename: str, user=Depends(require_auth)):
    safe_name = re.sub(r'[^\w.-]', '_', filename)
    if ".." in safe_name or "/" in safe_name:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")
    # Chercher dans le Nipogi d'abord, puis local
    for d in (STORAGE_DIR / "uploads" / "factures", DOCS_DIR / "factures"):
        filepath = d / safe_name
        if filepath.is_file():
            return FileResponse(str(filepath))
    raise HTTPException(status_code=404, detail="Facture non trouvee")

# ============ IMPORT EXCEL ============
@app.post("/api/import/excel")
async def import_excel(request: Request, user=Depends(require_referent)):
    """Importe un fichier Excel (.xlsx) et ajoute les contacts manquants.
    Détecte les doublons par (prenom+nom) ou telephone ou email.
    Les colonnes reconnues: Prenom, Nom, qualite, numero de telephone, email, Notes
    Les en-têtes sont comparées insensiblement à la casse et aux accents.
    """
    import openpyxl
    import unicodedata

    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier Excel requis")

    content = await file.read()
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Erreur lecture Excel: {str(e)}")
    
    ws = wb[wb.sheetnames[0]]
    
    # Lire les en-tetes (ligne 1)
    headers_raw = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers_raw.append(str(val).strip() if val else "")
    
    # Normaliser les en-tetes (minuscules, sans accents, sans espaces)
    def normalize_header(h):
        h = h.lower().strip()
        h = unicodedata.normalize('NFD', h).encode('ascii', 'ignore').decode()
        h = re.sub(r'[^a-z0-9]', '', h)
        return h
    
    headers_norm = [normalize_header(h) for h in headers_raw]
    
    # Mapper les colonnes - recherche flexible
    col_map = {}
    for i, h in enumerate(headers_norm):
        if h in ("prenom", "firstname", "first") or "prenom" in h:
            col_map["prenom"] = i + 1
        elif h in ("nom", "lastname", "name", "lastname") or "nom" in h and "prenom" not in h:
            col_map["nom"] = i + 1
        elif h in ("qualite", "role", "fonction", "qualite") or "qualite" in h:
            col_map["qualite"] = i + 1
        elif "telephone" in h or "tel" in h or "phone" in h or "mobile" in h:
            col_map["telephone"] = i + 1
        elif "email" in h or "mail" in h or "courriel" in h:
            col_map["email"] = i + 1
        elif "note" in h:
            col_map["notes"] = i + 1
    
    # Lire les donnees
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for field, col_num in col_map.items():
            val = ws.cell(row=row_idx, column=col_num).value
            row_data[field] = str(val).strip() if val else ""
        
        # Skip empty rows
        if not row_data.get("prenom") and not row_data.get("nom"):
            continue
        rows.append(row_data)
    
    os.unlink(tmp_path)
    
    if not rows:
        raise HTTPException(status_code=400, detail="Aucune ligne trouvee dans le fichier Excel")
    
    # Importer dans la base
    data = load_db()
    existing = data["contacts"]
    
    def find_duplicate(row):
        for c in existing:
            # Match par prenom+nom (insensible casse)
            if row.get("prenom", "").lower() == c.get("prenom", "").lower() and \
               row.get("nom", "").upper() == c.get("nom", "").upper():
                return c
            # Match par telephone normalise
            row_tel = normalize_phone(row.get("telephone", ""))
            if row_tel and row_tel == c.get("telephone", ""):
                return c
            # Match par email
            row_email = row.get("email", "").lower()
            if row_email and row_email == c.get("email", "").lower():
                return c
        return None
    
    added = []
    skipped = []
    
    for row in rows:
        dup = find_duplicate(row)
        if dup:
            skipped.append(f"{row.get('prenom','')} {row.get('nom','')}")
            continue
        
        contact = {
            "id": gen_id(),
            "prenom": row.get("prenom", "").strip(),
            "nom": row.get("nom", "").strip().upper(),
            "qualite": row.get("qualite", "").strip(),
            "telephone": normalize_phone(row.get("telephone", "")),
            "email": row.get("email", "").strip().lower(),
            "notes": row.get("notes", "").strip(),
            "historique": [],
            "cree_par": user["username"],
            "cree_le": now_iso(),
            "modifie_par": "",
            "modifie_le": ""
        }
        add_history(contact, user, "import_excel", f"Importe depuis Excel par {user['username']}")
        existing.append(contact)
        added.append(f"{contact['prenom']} {contact['nom']}")
    
    if added:
        save_db(data)
        backup_db()
    
    return {
        "status": "ok",
        "total_lignes": len(rows),
        "ajoutes": len(added),
        "doublons_ignores": len(skipped),
        "nouveaux": added,
        "deja_existants": skipped,
        "colonnes_detectees": {k: headers_raw[v-1] for k, v in col_map.items()}
    }



# ============ GROUPES DE DIFFUSION ============
@app.get("/api/groupes/diffusion")
async def get_groupes_diffusion(user=Depends(require_referent)):
    """Retourne les groupes de diffusion par qualite avec le nombre de contacts ayant un email."""
    data = load_db()
    contacts = data["contacts"]
    groupes = {}
    for c in contacts:
        q = c.get("qualite", "non specifie") or "non specifie"
        if q not in groupes:
            groupes[q] = {"qualite": q, "total": 0, "avec_email": 0, "contacts": []}
        groupes[q]["total"] += 1
        if c.get("email"):
            groupes[q]["avec_email"] += 1
            groupes[q]["contacts"].append({"id": c["id"], "nom": f"{c.get('prenom','')} {c.get('nom','')}".strip(), "email": c["email"]})
    return list(groupes.values())

@app.post("/api/groupes/diffusion/send")
async def send_groupe_diffusion(req: EmailSend, user=Depends(require_referent)):
    """Envoie un email a un groupe de diffusion (contact_ids pre-remplis par le frontend)."""
    data = load_db()
    recipients = list(req.recipients)
    for cid in req.contact_ids:
        for c in data["contacts"]:
            if c["id"] == cid and c.get("email"):
                recipients.append(c["email"])
    recipients = list(set(r for r in recipients if r))
    if not recipients:
        raise HTTPException(status_code=400, detail="Aucun destinataire avec email")
    try:
        send_email_smtp(recipients, req.subject, req.body, html=req.html)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur envoi email: {str(e)}")
    for cid in req.contact_ids:
        for c in data["contacts"]:
            if c["id"] == cid:
                add_history(c, user, "email_groupe", f"Sujet: {req.subject} | A: {len(recipients)} destinataires")
    save_db(data)
    return {"status": "ok", "sent_to": recipients, "count": len(recipients)}

# ============ DASHBOARD DES ACTIVITES ============
@app.get("/api/dashboard")
async def get_dashboard(user=Depends(require_referent)):
    """Tableau de bord synthetique: activites recentes, cotisations en retard, prochaines echeances, contacts non contactes."""
    data = load_db()
    contacts = data["contacts"]
    today = datetime.now()
    annee_courante = str(today.year)
    
    # 1. Activites recentes (historique des 7 derniers jours)
    activites_recentes = []
    for c in contacts:
        for h in c.get("historique", []):
            try:
                h_date = datetime.fromisoformat(h.get("date", ""))
                if (today - h_date).days <= 7:
                    activites_recentes.append({
                        "date": h["date"],
                        "user": h.get("user", ""),
                        "action": h.get("action", ""),
                        "contact": f"{c.get('prenom','')} {c.get('nom','')}".strip(),
                        "contact_id": c["id"],
                        "details": h.get("details", "")
                    })
            except Exception:
                continue
    activites_recentes.sort(key=lambda x: x["date"], reverse=True)
    activites_recentes = activites_recentes[:20]
    
    # 2. Cotisations en retard (annee courante, non payees)
    cotisations_retard = []
    contact_by_id = {c["id"]: c for c in contacts}
    for cot in data.get("cotisations", []):
        if cot.get("annee") == annee_courante and cot.get("statut") != "paye":
            c = contact_by_id.get(cot.get("contact_id"), {})
            cotisations_retard.append({
                "contact_nom": f"{c.get('prenom','')} {c.get('nom','')}".strip(),
                "contact_id": cot.get("contact_id"),
                "montant": cot.get("montant", 0),
                "statut": cot.get("statut", ""),
                "date_paiement": cot.get("date_paiement", "")
            })
    
    # 3. Contacts sans cotisation pour l'annee courante
    contacts_avec_cot = {cot.get("contact_id") for cot in data.get("cotisations", []) if cot.get("annee") == annee_courante}
    contacts_sans_cot = []
    for c in contacts:
        if c["id"] not in contacts_avec_cot:
            contacts_sans_cot.append({
                "nom": f"{c.get('prenom','')} {c.get('nom','')}".strip(),
                "id": c["id"],
                "email": c.get("email", ""),
                "qualite": c.get("qualite", "")
            })
    
    # 4. Taches a venir (echeance dans les 30 prochains jours, non terminees)
    taches_a_venir = []
    for t in data.get("taches", []):
        if t.get("statut") != "terminee" and t.get("echeance"):
            try:
                echeance = datetime.strptime(t["echeance"], "%Y-%m-%d")
                if 0 <= (echeance - today).days <= 30:
                    contact = contact_by_id.get(t.get("assigne_a", ""), {})
                    taches_a_venir.append({
                        "titre": t.get("titre", ""),
                        "echeance": t["echeance"],
                        "priorite": t.get("priorite", "normale"),
                        "contact_nom": f"{contact.get('prenom','')} {contact.get('nom','')}".strip(),
                        "jours_restants": (echeance - today).days
                    })
            except Exception:
                continue
    taches_a_venir.sort(key=lambda x: x["echeance"])
    
    # 5. Prochains evenements
    prochains_evenements = []
    for ev in data.get("evenements", []):
        try:
            ev_date = datetime.strptime(ev.get("date", ""), "%Y-%m-%d")
            if ev_date >= today:
                prochains_evenements.append({
                    "titre": ev.get("titre", ""),
                    "date": ev.get("date", ""),
                    "lieu": ev.get("lieu", ""),
                    "jours_restants": (ev_date - today).days
                })
        except Exception:
            continue
    prochains_evenements.sort(key=lambda x: x["date"])
    
    # 6. Contacts non contactes depuis plus de 90 jours
    contacts_non_contactes = []
    for c in contacts:
        derniere_activite = c.get("modifie_le", "") or c.get("cree_le", "")
        for h in c.get("historique", []):
            if h.get("date", "") > derniere_activite:
                derniere_activite = h["date"]
        if derniere_activite:
            try:
                d = datetime.fromisoformat(derniere_activite)
                if (today - d).days > 90:
                    contacts_non_contactes.append({
                        "nom": f"{c.get('prenom','')} {c.get('nom','')}".strip(),
                        "id": c["id"],
                        "derniere_activite": derniere_activite[:10],
                        "jours": (today - d).days
                    })
            except Exception:
                continue
    
    return {
        "activites_recentes": activites_recentes,
        "cotisations_retard": cotisations_retard,
        "contacts_sans_cotisation": contacts_sans_cot,
        "taches_a_venir": taches_a_venir,
        "prochains_evenements": prochains_evenements,
        "contacts_non_contactes": contacts_non_contactes,
        "total_contacts": len(contacts),
        "total_cotisations_annee": len([c for c in data.get("cotisations",[]) if c.get("annee") == annee_courante]),
        "total_taches_actives": len([t for t in data.get("taches",[]) if t.get("statut") != "terminee"])
    }

# ============ RELANCE COTISATIONS ============
@app.post("/api/cotisations/relance")
async def relance_cotisations(annee: str = "", user=Depends(require_referent)):
    """Envoie un email de relance aux contacts sans cotisation payee pour l'annee donnee (defaut: annee courante)."""
    data = load_db()
    annee_cible = annee or str(datetime.now().year)
    
    # Trouver les contacts sans cotisation payee
    contacts_avec_cot_payee = {
        cot.get("contact_id") for cot in data.get("cotisations", [])
        if cot.get("annee") == annee_cible and cot.get("statut") == "paye"
    }
    
    a_relancer = []
    for c in data["contacts"]:
        if c["id"] not in contacts_avec_cot_payee and c.get("email"):
            a_relancer.append(c)
    
    if not a_relancer:
        return {"status": "ok", "envoyes": 0, "message": "Aucun contact a relancer (tous a jour ou sans email)"}
    
    subject = f"Relance cotisation {annee_cible} - Les Ami(e)s de Romy"
    body = f"""Bonjour,

Nous vous contactons concernant votre cotisation annuelle {annee_cible} pour l association Les Ami(e)s de Romy.

Si vous avez deja regle votre cotisation, merci d ignorer cet email.
Sinon, nous vous invitons a regulariser votre situation.

Montant de la cotisation: a definir selon le barème de l association.
Mode de paiement: especes, cheque ou virement.

Pour toute question, n hesitez pas a nous contacter.

Cordialement,
L equipe des Ami(e)s de Romy"""
    
    recipients = [c["email"] for c in a_relancer]
    try:
        send_email_smtp(recipients, subject, body)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur envoi email: {str(e)}")
    
    for c in a_relancer:
        add_history(c, user, "relance_cotisation", f"Relance cotisation {annee_cible} envoyee")
    save_db(data)
    
    return {"status": "ok", "envoyes": len(recipients), "destinataires": [c["email"] for c in a_relancer]}


# ============ FEATURE 2: SUIVI ADHESIONS ============
@app.get("/api/adhesions/suivi")
async def suivi_adhesions(user=Depends(require_referent)):
    """Pour chaque contact: statut cotisation (a_jour/expiree/en_retard/sans), annee derniere cotisation, jours depuis expiration."""
    data = load_db()
    contacts = data.get("contacts", [])
    cotisations = data.get("cotisations", [])
    today = datetime.now()
    annee_courante = str(today.year)

    # Index: contact_id -> cotisation la plus recente (annee maximale, statut paye)
    derniere_cot = {}
    for cot in cotisations:
        cid = cot.get("contact_id", "")
        if not cid:
            continue
        annee = cot.get("annee", "")
        # On ne considere que les cotisations payees pour le suivi d'adhesion
        if cot.get("statut") != "paye":
            continue
        prev = derniere_cot.get(cid)
        if prev is None or (annee and annee > prev.get("annee", "")):
            derniere_cot[cid] = cot

    resultats = []
    for c in contacts:
        cid = c.get("id", "")
        nom = f"{c.get('prenom','')} {c.get('nom','')}".strip()
        email = c.get("email", "")
        qualite = c.get("qualite", "")

        cot = derniere_cot.get(cid)
        if cot is None:
            # Aucune cotisation payee enregistree
            resultats.append({
                "contact_id": cid,
                "contact_nom": nom,
                "email": email,
                "qualite": qualite,
                "statut": "sans",
                "annee_derniere_cotisation": None,
                "jours_depuis_expiration": None
            })
            continue

        annee_derniere = cot.get("annee", "")
        # L'adhesion expire le 31/12 de l'annee de cotisation + 1 an
        # Donc pour cotisation annee N, expiration = 31/12/(N+1)
        try:
            annee_int = int(annee_derniere)
            date_expiration = datetime(annee_int + 1, 12, 31)
            jours_exp = (today - date_expiration).days
        except (ValueError, TypeError):
            jours_exp = None

        # Determiner le statut
        # a_jour: derniere cotisation = annee courante ou jour d'expiration non atteint
        # expiree: expiration dans 30 jours ou moins (bientot)
        # en_retard: deja expiree
        # sans: aucune cotisation
        if jours_exp is not None:
            if jours_exp <= 0 and annee_derniere == annee_courante:
                # A jour pour l'annee courante
                statut = "a_jour"
            elif jours_exp <= 0:
                # Pas encore expiree (annee precedente mais pas encore en retard)
                # Verifier si l'expiration est dans moins de 30 jours (negatif = bientot)
                statut = "a_jour"
            elif jours_exp <= 30:
                # Expire depuis moins de 30 jours = expire bientot (en fait deja expire, mais recent)
                statut = "expiree"
            else:
                # Expire depuis plus de 30 jours
                statut = "en_retard"
        else:
            statut = "sans"

        resultats.append({
            "contact_id": cid,
            "contact_nom": nom,
            "email": email,
            "qualite": qualite,
            "statut": statut,
            "annee_derniere_cotisation": annee_derniere,
            "jours_depuis_expiration": jours_exp
        })

    # Trier: en_retard > expiree > sans > a_jour
    ordre = {"en_retard": 0, "expiree": 1, "sans": 2, "a_jour": 3}
    resultats.sort(key=lambda r: (ordre.get(r["statut"], 4), r.get("contact_nom", "")))

    # Statistiques resumees
    resume = {
        "total": len(resultats),
        "a_jour": sum(1 for r in resultats if r["statut"] == "a_jour"),
        "expiree": sum(1 for r in resultats if r["statut"] == "expiree"),
        "en_retard": sum(1 for r in resultats if r["statut"] == "en_retard"),
        "sans": sum(1 for r in resultats if r["statut"] == "sans")
    }

    return {"contacts": resultats, "resume": resume}


@app.post("/api/adhesions/relance-auto")
async def relance_auto_adhesions(user=Depends(require_referent)):
    """Envoie un email automatique aux adherents dont la cotisation expire dans 30 jours ou est deja expiree."""
    data = load_db()
    contacts = data.get("contacts", [])
    cotisations = data.get("cotisations", [])
    today = datetime.now()
    annee_courante = str(today.year)

    # Index: contact_id -> cotisation payee la plus recente
    derniere_cot = {}
    for cot in cotisations:
        cid = cot.get("contact_id", "")
        if not cid or cot.get("statut") != "paye":
            continue
        annee = cot.get("annee", "")
        prev = derniere_cot.get(cid)
        if prev is None or (annee and annee > prev.get("annee", "")):
            derniere_cot[cid] = cot

    a_relancer = []
    contacts_sans_cot = []

    for c in contacts:
        cid = c.get("id", "")
        email = c.get("email", "")
        if not email:
            continue

        cot = derniere_cot.get(cid)
        if cot is None:
            # Sans cotisation - relancer aussi
            contacts_sans_cot.append(c)
            a_relancer.append({"contact": c, "type": "sans", "annee": None, "jours_exp": None})
            continue

        annee_derniere = cot.get("annee", "")
        try:
            annee_int = int(annee_derniere)
            date_expiration = datetime(annee_int + 1, 12, 31)
            jours_exp = (today - date_expiration).days
        except (ValueError, TypeError):
            continue

        # Relancer si: expire dans 30 jours (jours_exp >= -30 et <= 0) OU deja expiree (jours_exp > 0)
        # jours_exp negatif = pas encore expiree; -30 a 0 = expire bientot; >0 = deja expiree
        if -30 <= jours_exp <= 0:
            a_relancer.append({"contact": c, "type": "expire_bientot", "annee": annee_derniere, "jours_exp": jours_exp})
        elif jours_exp > 0:
            a_relancer.append({"contact": c, "type": "expiree", "annee": annee_derniere, "jours_exp": jours_exp})

    if not a_relancer:
        return {"status": "ok", "envoyes": 0, "message": "Aucun adherent a relancer (tous a jour ou sans email)"}

    smtp = load_smtp()
    if not smtp.get("password"):
        raise HTTPException(status_code=400, detail="SMTP non configure (smtp_config.json)")

    envoyes = 0
    erreurs = []
    for item in a_relancer:
        c = item["contact"]
        email = c.get("email", "")
        if not email:
            continue
        nom = f"{c.get('prenom','')} {c.get('nom','')}".strip()

        if item["type"] == "expire_bientot":
            sujet = f"Votre cotisation arrive a echeance - Les Ami(e)s de Romy"
            corps = f"""Bonjour {nom},

Votre cotisation annuelle pour l association Les Ami(e)s de Romy arrive a echeance prochainement.

Derniere cotisation enregistree: {item['annee']}
Nous vous invitons a regulariser votre cotisation pour l annee en cours.

Montant de la cotisation: a definir selon le bareme de l association.
Mode de paiement: especes, cheque ou virement.

Pour toute question, n hesitez pas a nous contacter.

Cordialement,
L equipe des Ami(e)s de Romy"""
        elif item["type"] == "expiree":
            sujet = f"Relance cotisation - Les Ami(e)s de Romy"
            corps = f"""Bonjour {nom},

Votre cotisation annuelle pour l association Les Ami(e)s de Romy est expiree depuis {item['jours_exp']} jours.

Derniere cotisation enregistree: {item['annee']}
Nous vous invitons a regulariser votre situation des que possible.

Montant de la cotisation: a definir selon le bareme de l association.
Mode de paiement: especes, cheque ou virement.

Pour toute question, n hesitez pas a nous contacter.

Cordialement,
L equipe des Ami(e)s de Romy"""
        else:  # sans cotisation
            sujet = f"Cotisation annuelle - Les Ami(e)s de Romy"
            corps = f"""Bonjour {nom},

Nous vous contactons concernant votre cotisation annuelle pour l association Les Ami(e)s de Romy.

Aucune cotisation n est actuellement enregistree a votre nom.
Nous vous invitons a regulariser votre adhesion.

Montant de la cotisation: a definir selon le bareme de l association.
Mode de paiement: especes, cheque ou virement.

Pour toute question, n hesitez pas a nous contacter.

Cordialement,
L equipe des Ami(e)s de Romy"""

        try:
            send_email_smtp([email], sujet, corps)
            envoyes += 1
            add_history(c, user, "relance_adhesion", f"Relance auto ({item['type']}) envoyee a {email}")
        except Exception as e:
            erreurs.append({"email": email, "erreur": str(e)})

    save_db(data)

    return {
        "status": "ok",
        "envoyes": envoyes,
        "erreurs": erreurs,
        "total_a_relancer": len(a_relancer),
        "destinataires": [item["contact"].get("email", "") for item in a_relancer if item["contact"].get("email")]
    }


# ============ F1: TABLEAU DE BORD D'IMPACT ============
@app.get("/api/impact")
async def get_impact(user=Depends(require_referent)):
    """Statistiques d'impact pour la prévention des violences infantiles."""
    data = load_db()
    contacts = data.get("contacts", [])
    events = data.get("evenements", [])
    cots = data.get("cotisations", [])
    dons = data.get("dons", [])
    accomps = data.get("accompagnements", [])

    now = datetime.now()
    year = str(now.year)

    # Dons par mois (12 derniers mois)
    dons_par_mois = {}
    for d in dons:
        dt = d.get("date", "")
        if dt and len(dt) >= 7:
            dons_par_mois[dt[:7]] = dons_par_mois.get(dt[:7], 0) + d.get("montant", 0)

    # Cotisations par année
    cots_par_annee = {}
    for c in cots:
        an = c.get("annee", "")
        if an:
            cots_par_annee[an] = cots_par_annee.get(an, 0) + c.get("montant", 0)

    # Contacts par qualité
    par_qualite = {}
    for c in contacts:
        q = c.get("qualite", "sans qualite")
        par_qualite[q] = par_qualite.get(q, 0) + 1

    # Événements par mois
    events_par_mois = {}
    for e in events:
        dt = e.get("date", "")
        if dt and len(dt) >= 7:
            events_par_mois[dt[:7]] = events_par_mois.get(dt[:7], 0) + 1

    # Accompagnements actifs
    accomp_actifs = [a for a in accomps if a.get("statut") == "actif"]
    accomp_par_type = {}
    for a in accomps:
        t = a.get("type_suivi", "accompagnement")
        accomp_par_type[t] = accomp_par_type.get(t, 0) + 1

    # Présence totale aux événements
    presence_total = 0
    for e in events:
        p = e.get("presence", {})
        presence_total += sum(1 for v in p.values() if v == "present")

    return {
        "contacts_total": len(contacts),
        "contacts_par_qualite": par_qualite,
        "events_total": len(events),
        "events_par_mois": events_par_mois,
        "presence_total": presence_total,
        "cots_par_annee": cots_par_annee,
        "dons_total": sum(d.get("montant", 0) for d in dons),
        "dons_par_mois": dons_par_mois,
        "accomp_total": len(accomps),
        "accomp_actifs": len(accomp_actifs),
        "accomp_par_type": accomp_par_type,
        "benevoles_actifs": len([c for c in contacts if c.get("qualite", "") in ("referent", "benevole", "membre asso")]),
    }


# ============ F2: CARTOGRAPHIE INTERVENANTS/PARTENAIRES ============
@app.get("/api/cartographie")
async def get_cartographie(user=Depends(require_referent)):
    """Retourne les contacts avec code postal/ville pour cartographie Leaflet."""
    data = load_db()
    contacts = data.get("contacts", [])
    points = []
    for c in contacts:
        cp = c.get("code_postal") or c.get("cp") or ""
        ville = c.get("ville") or ""
        # Extract from notes if not available
        if not cp and c.get("notes", ""):
            import re
            m = re.search(r'\b(\d{5})\b', c.get("notes", ""))
            if m:
                cp = m.group(1)
        if cp or ville:
            points.append({
                "id": c["id"],
                "nom": f"{c.get('prenom','')} {c.get('nom','')}",
                "qualite": c.get("qualite", ""),
                "email": c.get("email", ""),
                "telephone": c.get("telephone", ""),
                "code_postal": cp,
                "ville": ville,
                "notes": c.get("notes", "")[:100]
            })
    return {"points": points, "total": len(points)}


# ============ F3: SUIVI D'ACCOMPAGNEMENT FAMILIAL ============
@app.get("/api/accompagnements")
async def list_accompagnements(statut: str = "", user=Depends(require_referent)):
    data = load_db()
    accomps = data.get("accompagnements", [])
    if statut:
        accomps = [a for a in accomps if a.get("statut") == statut]
    # Enrich with contact name
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for a in accomps:
        c = contact_by_id.get(a["contact_id"])
        a["contact_nom"] = f"{c.get('prenom','')} {c.get('nom','')}" if c else "?"
        # Hide confidential notes from non-admin
        if user["role"] != "admin":
            a.pop("notes_confidentielles", None)
    return accomps

@app.post("/api/accompagnements")
async def create_accompagnement(req: AccompagnementCreate, user=Depends(require_referent)):
    data = load_db()
    if "accompagnements" not in data:
        data["accompagnements"] = []
    a = {
        "id": f"acc-{secrets.token_hex(8)}",
        "contact_id": req.contact_id,
        "type_suivi": req.type_suivi,
        "date_debut": req.date_debut,
        "date_fin": req.date_fin,
        "statut": req.statut,
        "priorite": req.priorite,
        "intervenants": req.intervenants,
        "notes_confidentielles": req.notes_confidentielles,
        "notes_partagees": req.notes_partagees,
        "historique": [{"date": now_iso(), "user": user["username"], "action": "creation"}],
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["accompagnements"].append(a)
    save_db(data)
    backup_db()
    return {"status": "ok", "accompagnement": a}

@app.put("/api/accompagnements/{acc_id}")
async def update_accompagnement(acc_id: str, req: AccompagnementUpdate, user=Depends(require_referent)):
    data = load_db()
    for a in data.get("accompagnements", []):
        if a["id"] == acc_id:
            changes = []
            for field in ["type_suivi", "date_debut", "date_fin", "statut", "priorite", "intervenants", "notes_partagees"]:
                val = getattr(req, field)
                if val is not None and a.get(field) != val:
                    changes.append(f"{field} modifie")
                    a[field] = val
            # Confidential notes only for admin
            if req.notes_confidentielles is not None and user["role"] == "admin":
                a["notes_confidentielles"] = req.notes_confidentielles
                changes.append("notes confidentielles modifiees")
            if changes:
                a["historique"].append({"date": now_iso(), "user": user["username"], "action": "modification", "details": " | ".join(changes)})
                save_db(data)
                backup_db()
            return {"status": "ok", "accompagnement": a}
    raise HTTPException(status_code=404, detail="Accompagnement non trouve")

@app.post("/api/accompagnements/{acc_id}/notes")
async def add_accompagnement_note(acc_id: str, req: AccompagnementNote, user=Depends(require_referent)):
    data = load_db()
    for a in data.get("accompagnements", []):
        if a["id"] == acc_id:
            entry = {
                "date": now_iso(),
                "user": user["username"],
                "note": req.note,
                "confidentiel": req.confidentiel and user["role"] == "admin"
            }
            a.setdefault("notes_historique", []).append(entry)
            a["historique"].append({"date": now_iso(), "user": user["username"], "action": "note", "details": req.note[:80]})
            save_db(data)
            backup_db()
            return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Accompagnement non trouve")

@app.delete("/api/accompagnements/{acc_id}")
async def delete_accompagnement(acc_id: str, user=Depends(require_referent)):
    data = load_db()
    data["accompagnements"] = [a for a in data.get("accompagnements", []) if a["id"] != acc_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}


# ============ F4: PLANNING PRESENCE BENEVOLS ============
@app.get("/api/planning")
async def get_planning(evenement_id: str = "", user=Depends(require_auth)):
    data = load_db()
    planning = data.get("planning_benevoles", [])
    if evenement_id:
        planning = [p for p in planning if p["evenement_id"] == evenement_id]
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for p in planning:
        c = contact_by_id.get(p["contact_id"])
        p["contact_nom"] = f"{c.get('prenom','')} {c.get('nom','')}" if c else "?"
    return planning

@app.post("/api/planning")
async def set_planning(req: BenevolePlanning, user=Depends(require_referent)):
    data = load_db()
    if "planning_benevoles" not in data:
        data["planning_benevoles"] = []
    # Remove existing entry for same event+contact, then add
    data["planning_benevoles"] = [p for p in data["planning_benevoles"]
                                  if not (p["evenement_id"] == req.evenement_id and p["contact_id"] == req.contact_id)]
    entry = {
        "id": f"plan-{secrets.token_hex(8)}",
        "evenement_id": req.evenement_id,
        "contact_id": req.contact_id,
        "role": req.role,
        "creneau": req.creneau,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["planning_benevoles"].append(entry)
    save_db(data)
    backup_db()
    return {"status": "ok", "planning": entry}

@app.delete("/api/planning/{plan_id}")
async def delete_planning(plan_id: str, user=Depends(require_referent)):
    data = load_db()
    data["planning_benevoles"] = [p for p in data.get("planning_benevoles", []) if p["id"] != plan_id]
    save_db(data)
    backup_db()
    return {"status": "ok"}


# ============ F5: GENERATEUR RAPPORTS D'ACTIVITE ============
@app.get("/api/rapport-activite")
async def generate_rapport(annee: str = "", token: str = ""):
    """Genere un rapport d'activite au format HTML (pour impression PDF navigateur). Token en query car window.open ne peut pas envoyer de header."""
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token invalide")
    if user["role"] not in REFERENT_ROLES:
        raise HTTPException(status_code=403, detail="Acces referent requis")
    data = load_db()
    annee_cible = annee or str(datetime.now().year)
    contacts = data.get("contacts", [])
    events = data.get("evenements", [])
    cots = data.get("cotisations", [])
    dons = data.get("dons", [])
    accomps = data.get("accompagnements", [])

    # Filter by year
    events_an = [e for e in events if e.get("date", "").startswith(annee_cible)]
    cots_an = [c for c in cots if c.get("annee") == annee_cible]
    dons_an = [d for d in dons if d.get("date", "").startswith(annee_cible)]
    accomp_an = [a for a in accomps if a.get("date_debut", "").startswith(annee_cible)]

    total_cots = sum(c.get("montant", 0) for c in cots_an)
    total_dons = sum(d.get("montant", 0) for d in dons_an)

    # Presence stats
    presence_total = sum(sum(1 for v in e.get("presence", {}).values() if v == "present") for e in events_an)

    html = f"""<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8"><title>Rapport d activite {annee_cible}</title>
<style>
body{{font-family:Arial,sans-serif;max-width:800px;margin:0 auto;padding:2rem;color:#333}}
h1{{color:#d03ec6;text-align:center;border-bottom:3px solid #d03ec6;padding-bottom:.5rem}}
h2{{color:#9e2b94;margin-top:2rem}}
.stat-box{{display:inline-block;width:45%;margin:1%;padding:1rem;background:#f3e0f5;border-radius:10px;text-align:center}}
.stat-box .num{{font-size:2rem;font-weight:bold;color:#d03ec6}}
.stat-box .lbl{{font-size:.85rem;color:#666}}
table{{width:100%;border-collapse:collapse;margin:1rem 0}}
th{{background:#d03ec6;color:#fff;padding:.5rem;text-align:left}}
td{{padding:.5rem;border-bottom:1px solid #eee}}
.footer{{margin-top:3rem;text-align:center;font-size:.8rem;color:#999}}
</style></head><body>
<h1>Rapport d activite {annee_cible}</h1>
<p style="text-align:center;color:#666">Les Ami(e)s de Romy — Prevention des violences infantiles</p>

<h2>Chiffres cles</h2>
<div class="stat-box"><div class="num">{len(contacts)}</div><div class="lbl">Contacts total</div></div>
<div class="stat-box"><div class="num">{len(events_an)}</div><div class="lbl">Evenements {annee_cible}</div></div>
<div class="stat-box"><div class="num">{presence_total}</div><div class="lbl">Presences aux evenements</div></div>
<div class="stat-box"><div class="num">{len(accomp_an)}</div><div class="lbl">Accompagnements</div></div>
<div class="stat-box"><div class="num">{total_cots:.0f} &euro;</div><div class="lbl">Cotisations {annee_cible}</div></div>
<div class="stat-box"><div class="num">{total_dons:.0f} &euro;</div><div class="lbl">Dons {annee_cible}</div></div>

<h2>Evenements de l annee</h2>
<table><tr><th>Date</th><th>Titre</th><th>Lieu</th><th>Presences</th></tr>"""
    for e in events_an:
        pres = sum(1 for v in e.get("presence", {}).values() if v == "present")
        html += f"<tr><td>{e.get('date','')}</td><td>{e.get('titre','')}</td><td>{e.get('lieu','')}</td><td>{pres}</td></tr>"
    html += "</table>"

    html += f"<h2>Dons de l annee</h2><table><tr><th>Date</th><th>Nom</th><th>Montant</th></tr>"
    for d in dons_an:
        html += f"<tr><td>{d.get('date','')}</td><td>{d.get('nom','')}</td><td>{d.get('montant',0)} &euro;</td></tr>"
    html += "</table>"

    html += f"<h2>Accompagnements {annee_cible}</h2><table><tr><th>Type</th><th>Statut</th><th>Priorite</th></tr>"
    for a in accomp_an:
        html += f"<tr><td>{a.get('type_suivi','')}</td><td>{a.get('statut','')}</td><td>{a.get('priorite','')}</td></tr>"
    html += "</table>"

    html += f'<div class="footer">Genere le {datetime.now().strftime("%d/%m/%Y")} par {user["username"]} — CRM Les Ami(e)s de Romy v{VERSION}</div>'
    html += "</body></html>"

    return HTMLResponse(content=html)


# ============ F6: NOTES DE FRAIS ============
@app.get("/api/notes-frais")
async def list_notes_frais(statut: str = "", user=Depends(require_referent)):
    data = load_db()
    notes = data.get("notes_frais", [])
    if statut:
        notes = [n for n in notes if n.get("statut") == statut]
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for n in notes:
        c = contact_by_id.get(n.get("contact_id", ""))
        n["contact_nom"] = f"{c.get('prenom','')} {c.get('nom','')}" if c else n.get("contact_nom", "-")
    return notes

@app.post("/api/notes-frais")
async def create_note_frais(req: NoteFraisCreate, user=Depends(require_referent)):
    data = load_db()
    if "notes_frais" not in data:
        data["notes_frais"] = []
    n = {
        "id": f"nf-{secrets.token_hex(8)}",
        "contact_id": req.contact_id,
        "date": req.date,
        "montant": req.montant,
        "categorie": req.categorie,
        "description": req.description,
        "justificatif": req.justificatif,
        "statut": req.statut,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["notes_frais"].append(n)
    save_db(data)
    backup_db()
    return {"status": "ok", "note_frais": n}

@app.put("/api/notes-frais/{nf_id}")
async def update_note_frais(nf_id: str, req: NoteFraisUpdate, user=Depends(require_referent)):
    data = load_db()
    for n in data.get("notes_frais", []):
        if n["id"] == nf_id:
            if req.montant is not None: n["montant"] = req.montant
            if req.categorie is not None: n["categorie"] = req.categorie
            if req.description is not None: n["description"] = req.description
            if req.statut is not None: n["statut"] = req.statut
            save_db(data); backup_db()
            return {"status": "ok", "note_frais": n}
    raise HTTPException(status_code=404, detail="Note de frais non trouvee")

@app.delete("/api/notes-frais/{nf_id}")
async def delete_note_frais(nf_id: str, user=Depends(require_referent)):
    data = load_db()
    data["notes_frais"] = [n for n in data.get("notes_frais", []) if n["id"] != nf_id]
    save_db(data); backup_db()
    return {"status": "ok"}

@app.get("/api/notes-frais/total")
async def total_notes_frais(user=Depends(require_referent)):
    data = load_db()
    notes = data.get("notes_frais", [])
    return {
        "total": sum(n.get("montant", 0) for n in notes),
        "en_attente": sum(n.get("montant", 0) for n in notes if n.get("statut") == "en_attente"),
        "valide": sum(n.get("montant", 0) for n in notes if n.get("statut") == "valide"),
        "rembourse": sum(n.get("montant", 0) for n in notes if n.get("statut") == "rembourse"),
        "count": len(notes)
    }


# ---- OCR factures via Tesseract (local, sans Ollama) ----
# Les factures sont stockées dans uploads/factures/
_NIPOGI_FACTURES = STORAGE_DIR / "uploads" / "factures"
_LOCAL_FACTURES = DOCS_DIR / "factures"
FACTURES_DIR = _NIPOGI_FACTURES if _NIPOGI_FACTURES.is_dir() else _LOCAL_FACTURES
try:
    FACTURES_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    FACTURES_DIR = _LOCAL_FACTURES
    FACTURES_DIR.mkdir(exist_ok=True)

# Tesseract OCR — installé dans le conteneur via apt + pip
try:
    import pytesseract
    from PIL import Image as PILImage
    _TESSERACT_OK = True
except ImportError:
    _TESSERACT_OK = False


def _extract_ocr_data(image_bytes: bytes) -> dict:
    """Extract invoice data from image using Tesseract OCR + regex parsing."""
    if not _TESSERACT_OK:
        raise RuntimeError("Tesseract non installe")
    
    import io as _io
    import re as _re
    
    img = PILImage.open(_io.BytesIO(image_bytes))
    # Run OCR with French language
    raw_text = pytesseract.image_to_string(img, lang="fra+eng")
    
    # Extract data with regex
    result = {"montant": 0.0, "date": "", "fournisseur": "", "categorie": "autre", "description": ""}
    
    # Montant: look for patterns like "Total: 42,50 €" or "42.50 EUR" or "Total TTC: 25,50"
    montant_patterns = [
        r'(?:total\s*(?:ttc)?\s*[:\-]?\s*)(\d+[.,]\d{1,2})\s*(?:€|EUR|euro)?',
        r'(\d+[.,]\d{2})\s*(?:€|EUR|euro)',
        r'montant\s*(?:total\s*)?[:\-]?\s*(\d+[.,]\d{1,2})',
    ]
    for pat in montant_patterns:
        m = _re.search(pat, raw_text, _re.IGNORECASE)
        if m:
            val = m.group(1).replace(",", ".")
            try:
                result["montant"] = float(val)
                break
            except ValueError:
                pass
    
    # Date: look for DD/MM/YYYY or DD-MM-YYYY or YYYY-MM-DD
    date_patterns = [
        r'(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})',
        r'(\d{4}-\d{2}-\d{2})',
        r'date\s*[:\-]?\s*(\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4})',
    ]
    for pat in date_patterns:
        m = _re.search(pat, raw_text, _re.IGNORECASE)
        if m:
            date_str = m.group(1)
            # Normalize to YYYY-MM-DD
            try:
                parts = _re.split(r'[/.-]', date_str)
                if len(parts[0]) == 4:  # YYYY-MM-DD
                    result["date"] = date_str
                elif len(parts[2]) == 4:  # DD/MM/YYYY
                    result["date"] = f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
                else:  # DD/MM/YY
                    year = parts[2]
                    if len(year) == 2:
                        year = "20" + year
                    result["date"] = f"{year}-{int(parts[1]):02d}-{int(parts[0]):02d}"
                break
            except (ValueError, IndexError):
                pass
    
    # Fournisseur: first non-empty line that's not a date or "facture"
    lines = [l.strip() for l in raw_text.split("\n") if l.strip()]
    skip_words = {"facture", "ticket", "total", "date", "montant", "tva", "€", "eur"}
    for line in lines[:10]:
        lower = line.lower()
        if not any(w in lower for w in skip_words) and not _re.match(r'^\d', line) and len(line) > 2:
            result["fournisseur"] = line[:100]
            break
    
    # Categorie: guess from keywords
    text_lower = raw_text.lower()
    if any(w in text_lower for w in ["repas", "restaurant", "cafe", "boulanger", "pizzeria", "mcdonald", "bar"]):
        result["categorie"] = "repas"
    elif any(w in text_lower for w in ["essence", "train", "bus", "metro", "taxi", "parking", "peage", "deplacement", "km"]):
        result["categorie"] = "deplacement"
    elif any(w in text_lower for w in ["fourniture", "papeterie", "bureau", "stylo", "classeur"]):
        result["categorie"] = "fournitures"
    elif any(w in text_lower for w in ["telephone", "internet", "mobile", "forfait", "communication"]):
        result["categorie"] = "communication"
    
    # Description: first meaningful line after fournisseur
    result["description"] = (result["fournisseur"] or "Facture")[:100]
    
    return result


@app.post("/api/frais/ocr")
async def ocr_facture(request: Request, user=Depends(require_referent)):
    """Upload a facture image, save it, and run OCR via Ollama to extract data."""
    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier image requis")

    ext = Path(file.filename).suffix.lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
        raise HTTPException(status_code=400, detail="Format image requis (jpg, png, gif, webp)")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image trop volumineuse (max 20 Mo)")

    safe_name = re.sub(r'[^\w.-]', '_', file.filename).replace("..", "").replace("/", "").replace("\\", "")
    if not safe_name:
        safe_name = f"facture_{secrets.token_hex(4)}{ext}"
    dest = FACTURES_DIR / safe_name
    if dest.exists():
        stem = Path(safe_name).stem
        dest = FACTURES_DIR / f"{stem}_{secrets.token_hex(4)}{ext}"
        safe_name = dest.name
    with open(dest, "wb") as f:
        f.write(content)

    try:
        ocr_data = _extract_ocr_data(content)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=503, detail=f"OCR indisponible: {e}")

    ocr_data["image_path"] = safe_name
    return ocr_data


@app.post("/api/frais/scan-and-create")
async def scan_and_create_note_frais(request: Request, user=Depends(require_referent)):
    """Upload image, OCR it, and create a note de frais in one step."""
    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier image requis")

    ext = Path(file.filename).suffix.lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
        raise HTTPException(status_code=400, detail="Format image requis (jpg, png, gif, webp)")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image trop volumineuse (max 20 Mo)")

    safe_name = re.sub(r'[^\w.-]', '_', file.filename).replace("..", "").replace("/", "").replace("\\", "")
    if not safe_name:
        safe_name = f"facture_{secrets.token_hex(4)}{ext}"
    dest = FACTURES_DIR / safe_name
    if dest.exists():
        stem = Path(safe_name).stem
        dest = FACTURES_DIR / f"{stem}_{secrets.token_hex(4)}{ext}"
        safe_name = dest.name
    with open(dest, "wb") as f:
        f.write(content)

    try:
        ocr_data = _extract_ocr_data(content)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=503, detail=f"OCR indisponible: {e}")

    ocr_data["image_path"] = safe_name

    # Create note de frais
    data = load_db()
    if "notes_frais" not in data:
        data["notes_frais"] = []
    n = {
        "id": f"nf-{secrets.token_hex(8)}",
        "contact_id": "",
        "date": ocr_data.get("date", ""),
        "montant": ocr_data.get("montant", 0),
        "categorie": ocr_data.get("categorie", "autre"),
        "description": ocr_data.get("description", ""),
        "justificatif": safe_name,
        "statut": "en_attente",
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["notes_frais"].append(n)
    save_db(data)
    backup_db()

    return {"status": "ok", "note_frais": n, "ocr": ocr_data}


# ============ F7: ECHEANCES & RAPPELS ============
@app.get("/api/echeances")
async def list_echeances(statut: str = "", user=Depends(require_referent)):
    data = load_db()
    echeances = data.get("echeances", [])
    # Auto-update overdue status
    today = datetime.now().strftime("%Y-%m-%d")
    for e in echeances:
        if e.get("statut") == "a_venir" and e.get("date") and e["date"] < today:
            e["statut"] = "en_retard"
    if statut:
        echeances = [e for e in echeances if e.get("statut") == statut]
    save_db(data)  # persist status changes
    return echeances

@app.post("/api/echeances")
async def create_echeance(req: EcheanceCreate, user=Depends(require_referent)):
    data = load_db()
    if "echeances" not in data:
        data["echeances"] = []
    e = {
        "id": f"ech-{secrets.token_hex(8)}",
        "titre": req.titre,
        "date": req.date,
        "type_echeance": req.type_echeance,
        "description": req.description,
        "recursif": req.recursif,
        "responsable": req.responsable,
        "statut": "a_venir",
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["echeances"].append(e)
    save_db(data); backup_db()
    return {"status": "ok", "echeance": e}

@app.put("/api/echeances/{ech_id}")
async def update_echeance(ech_id: str, req: EcheanceUpdate, user=Depends(require_referent)):
    data = load_db()
    for e in data.get("echeances", []):
        if e["id"] == ech_id:
            for f in ["titre", "date", "type_echeance", "description", "recursif", "responsable", "statut"]:
                val = getattr(req, f)
                if val is not None: e[f] = val
            # If marked done and recursive, create next occurrence
            if req.statut == "fait" and e.get("recursif") and e.get("date"):
                d = datetime.strptime(e["date"], "%Y-%m-%d")
                if e["recursif"] == "annuel":
                    d = d.replace(year=d.year + 1)
                elif e["recursif"] == "mensuel":
                    # stdlib month increment
                    m = d.month + 1
                    y = d.year
                    if m > 12: m = 1; y += 1
                    import calendar
                    last_day = calendar.monthrange(y, m)[1]
                    d = d.replace(year=y, month=m, day=min(d.day, last_day))
                new_e = dict(e)
                new_e["id"] = f"ech-{secrets.token_hex(8)}"
                new_e["date"] = d.strftime("%Y-%m-%d")
                new_e["statut"] = "a_venir"
                new_e["cree_le"] = now_iso()
                data["echeances"].append(new_e)
            save_db(data); backup_db()
            return {"status": "ok", "echeance": e}
    raise HTTPException(status_code=404, detail="Echeance non trouvee")

@app.delete("/api/echeances/{ech_id}")
async def delete_echeance(ech_id: str, user=Depends(require_referent)):
    data = load_db()
    data["echeances"] = [e for e in data.get("echeances", []) if e["id"] != ech_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F8: REGISTRE DES APPELS ============
@app.get("/api/registre-appels")
async def list_registre_appels(motif: str = "", user=Depends(require_referent)):
    data = load_db()
    appels = data.get("registre_appels", [])
    if motif:
        appels = [a for a in appels if a.get("motif") == motif]
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for a in appels:
        c = contact_by_id.get(a.get("contact_id", ""))
        a["contact_nom"] = f"{c.get('prenom','')} {c.get('nom','')}" if c else ""
    return appels

@app.post("/api/registre-appels")
async def create_registre_appel(req: RegistreAppelCreate, user=Depends(require_referent)):
    data = load_db()
    if "registre_appels" not in data:
        data["registre_appels"] = []
    a = {
        "id": f"app-{secrets.token_hex(8)}",
        "contact_id": req.contact_id,
        "nom_appelant": req.nom_appelant,
        "nom_appele": req.nom_appele,
        "date_appel": req.date_appel,
        "heure": req.heure,
        "motif": req.motif,
        "description": req.description,
        "suite_donnee": req.suite_donnee,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["registre_appels"].append(a)
    save_db(data); backup_db()
    return {"status": "ok", "appel": a}

@app.delete("/api/registre-appels/{appel_id}")
async def delete_registre_appel(appel_id: str, user=Depends(require_referent)):
    data = load_db()
    data["registre_appels"] = [a for a in data.get("registre_appels", []) if a["id"] != appel_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F9: VOTES / SCRUTINS AG ============
@app.get("/api/votes")
async def list_votes(user=Depends(require_referent)):
    data = load_db()
    return data.get("votes", [])

@app.post("/api/votes")
async def create_vote(req: VoteCreate, user=Depends(require_referent)):
    data = load_db()
    if "votes" not in data:
        data["votes"] = []
    v = {
        "id": f"vote-{secrets.token_hex(8)}",
        "titre": req.titre,
        "date": req.date,
        "type_vote": req.type_vote,
        "description": req.description,
        "resultat": {"pour": 0, "contre": 0, "abstention": 0},
        "notes": "",
        "statut": "ouvert",  # ouvert, cloture
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["votes"].append(v)
    save_db(data); backup_db()
    return {"status": "ok", "vote": v}

@app.put("/api/votes/{vote_id}")
async def update_vote_resultat(vote_id: str, req: VoteResultat, user=Depends(require_referent)):
    data = load_db()
    for v in data.get("votes", []):
        if v["id"] == vote_id:
            v["resultat"] = {"pour": req.pour, "contre": req.contre, "abstention": req.abstention}
            v["notes"] = req.notes
            v["statut"] = "cloture"
            save_db(data); backup_db()
            return {"status": "ok", "vote": v}
    raise HTTPException(status_code=404, detail="Vote non trouve")

@app.delete("/api/votes/{vote_id}")
async def delete_vote(vote_id: str, user=Depends(require_referent)):
    data = load_db()
    data["votes"] = [v for v in data.get("votes", []) if v["id"] != vote_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F10: CHECKLISTS RECURRENTES ============
@app.get("/api/checklists")
async def list_checklists(type_checklist: str = "", user=Depends(require_referent)):
    data = load_db()
    cls = data.get("checklists", [])
    if type_checklist:
        cls = [c for c in cls if c.get("type_checklist") == type_checklist]
    return cls

@app.post("/api/checklists")
async def create_checklist(req: ChecklistCreate, user=Depends(require_referent)):
    data = load_db()
    if "checklists" not in data:
        data["checklists"] = []
    c = {
        "id": f"chk-{secrets.token_hex(8)}",
        "titre": req.titre,
        "type_checklist": req.type_checklist,
        "evenement_id": req.evenement_id,
        "taches": req.taches,
        "coche": [False] * len(req.taches),
        "statut": "en_cours",
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["checklists"].append(c)
    save_db(data); backup_db()
    return {"status": "ok", "checklist": c}

@app.put("/api/checklists/{chk_id}")
async def update_checklist(chk_id: str, req: ChecklistUpdate, user=Depends(require_referent)):
    data = load_db()
    for c in data.get("checklists", []):
        if c["id"] == chk_id:
            if req.taches is not None:
                c["taches"] = req.taches
                c["coche"] = req.coche or [False] * len(req.taches)
            elif req.coche is not None:
                c["coche"] = req.coche
            # Check if all done
            if all(c["coche"]):
                c["statut"] = "termine"
            save_db(data); backup_db()
            return {"status": "ok", "checklist": c}
    raise HTTPException(status_code=404, detail="Checklist non trouvee")

@app.delete("/api/checklists/{chk_id}")
async def delete_checklist(chk_id: str, user=Depends(require_referent)):
    data = load_db()
    data["checklists"] = [c for c in data.get("checklists", []) if c["id"] != chk_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F11: SONDAGES RAPIDES ============
@app.get("/api/sondages")
async def list_sondages(user=Depends(require_auth)):
    data = load_db()
    sondages = data.get("sondages", [])
    # Members see only sondages targeting them
    if user["role"] == "membre":
        sondages = [s for s in sondages if s.get("cible") in ("membres", "tous") and s.get("statut") == "ouvert"]
    # Add response counts
    for s in sondages:
        s["nb_repondants"] = len(s.get("reponses", {}))
        s["resultats"] = {}
        for r in s.get("reponses", {}).values():
            for idx in (r if isinstance(r, list) else [r]):
                s["resultats"][idx] = s["resultats"].get(idx, 0) + 1
    return sondages

@app.post("/api/sondages")
async def create_sondage(req: SondageCreate, user=Depends(require_referent)):
    data = load_db()
    if "sondages" not in data:
        data["sondages"] = []
    s = {
        "id": f"snd-{secrets.token_hex(8)}",
        "question": req.question,
        "options": req.options,
        "cible": req.cible,
        "date_fin": req.date_fin,
        "multiple": req.multiple,
        "statut": "ouvert",
        "reponses": {},  # {contact_id: option_index or [indexes]}
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["sondages"].append(s)
    save_db(data); backup_db()
    return {"status": "ok", "sondage": s}

@app.post("/api/sondages/{sondage_id}/repondre")
async def repondre_sondage(sondage_id: str, req: SondageReponse, user=Depends(require_auth)):
    data = load_db()
    for s in data.get("sondages", []):
        if s["id"] == sondage_id and s.get("statut") == "ouvert":
            cid = user.get("contact_id") or user["username"]
            if s.get("multiple"):
                # Toggle option
                current = s["reponses"].get(cid, [])
                if not isinstance(current, list): current = [current]
                if req.option_index in current:
                    current = [x for x in current if x != req.option_index]
                else:
                    current.append(req.option_index)
                s["reponses"][cid] = current
            else:
                s["reponses"][cid] = req.option_index
            save_db(data); backup_db()
            return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Sondage non trouve ou cloture")

@app.put("/api/sondages/{sondage_id}/cloturer")
async def cloturer_sondage(sondage_id: str, user=Depends(require_referent)):
    data = load_db()
    for s in data.get("sondages", []):
        if s["id"] == sondage_id:
            s["statut"] = "cloture"
            save_db(data); backup_db()
            return {"status": "ok"}
    raise HTTPException(status_code=404, detail="Sondage non trouve")

@app.delete("/api/sondages/{sondage_id}")
async def delete_sondage(sondage_id: str, user=Depends(require_referent)):
    data = load_db()
    data["sondages"] = [s for s in data.get("sondages", []) if s["id"] != sondage_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F12: COMPTES-RENDUS DE REUNIONS ============
@app.get("/api/comptes-rendus")
async def list_comptes_rendus(user=Depends(require_auth)):
    data = load_db()
    crs = data.get("comptes_rendus", [])
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for cr in crs:
        cr["presents_noms"] = [f"{contact_by_id.get(cid,{}).get('prenom','')} {contact_by_id.get(cid,{}).get('nom','')}" for cid in cr.get("presents", [])]
        cr["nb_presents"] = len(cr.get("presents", []))
    return crs

@app.get("/api/comptes-rendus/{cr_id}")
async def get_compte_rendu(cr_id: str, user=Depends(require_auth)):
    data = load_db()
    contact_by_id = {c["id"]: c for c in data["contacts"]}
    for cr in data.get("comptes_rendus", []):
        if cr["id"] == cr_id:
            cr["presents_noms"] = [f"{contact_by_id.get(cid,{}).get('prenom','')} {contact_by_id.get(cid,{}).get('nom','')}" for cid in cr.get("presents", [])]
            cr["absents_noms"] = [f"{contact_by_id.get(cid,{}).get('prenom','')} {contact_by_id.get(cid,{}).get('nom','')}" for cid in cr.get("absents", [])]
            cr["excused_noms"] = [f"{contact_by_id.get(cid,{}).get('prenom','')} {contact_by_id.get(cid,{}).get('nom','')}" for cid in cr.get("excused", [])]
            return cr
    raise HTTPException(status_code=404, detail="Compte-rendu non trouve")

@app.post("/api/comptes-rendus")
async def create_compte_rendu(req: CompteRenduCreate, user=Depends(require_referent)):
    data = load_db()
    if "comptes_rendus" not in data:
        data["comptes_rendus"] = []
    cr = {
        "id": f"cr-{secrets.token_hex(8)}",
        "titre": req.titre, "date": req.date, "type_reunion": req.type_reunion,
        "presents": req.presents, "absents": req.absents, "excused": req.excused,
        "ordre_du_jour": req.ordre_du_jour, "discussions": req.discussions,
        "decisions": req.decisions, "actions": req.actions,
        "cree_par": user["username"], "cree_le": now_iso()
    }
    data["comptes_rendus"].append(cr)
    save_db(data); backup_db()
    return {"status": "ok", "compte_rendu": cr}

@app.put("/api/comptes-rendus/{cr_id}")
async def update_compte_rendu(cr_id: str, req: CompteRenduUpdate, user=Depends(require_referent)):
    data = load_db()
    for cr in data.get("comptes_rendus", []):
        if cr["id"] == cr_id:
            for f in ["titre","date","type_reunion","presents","absents","excused","ordre_du_jour","discussions","decisions","actions"]:
                val = getattr(req, f)
                if val is not None: cr[f] = val
            save_db(data); backup_db()
            return {"status": "ok", "compte_rendu": cr}
    raise HTTPException(status_code=404, detail="Compte-rendu non trouve")

@app.delete("/api/comptes-rendus/{cr_id}")
async def delete_compte_rendu(cr_id: str, user=Depends(require_referent)):
    data = load_db()
    data["comptes_rendus"] = [cr for cr in data.get("comptes_rendus", []) if cr["id"] != cr_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F12b: GENERATEUR DE PV DE REUNION ============
@app.get("/api/cr/{cr_id}/pv")
async def generate_pv(cr_id: str, token: str = ""):
    """Genere un PV de reunion au format HTML imprimable. Token en query car window.open ne peut pas envoyer de header."""
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token invalide")
    data = load_db()
    contact_by_id = {c["id"]: c for c in data.get("contacts", [])}

    # Find the CR
    cr = None
    for c in data.get("comptes_rendus", []):
        if c["id"] == cr_id:
            cr = c
            break
    if not cr:
        raise HTTPException(status_code=404, detail="Compte-rendu non trouve")

    # Helper to resolve contact names
    def noms(ids):
        return [f"{contact_by_id.get(cid, {}).get('prenom', '')} {contact_by_id.get(cid, {}).get('nom', '')}".strip()
                for cid in (ids or []) if contact_by_id.get(cid)]

    presents_noms = noms(cr.get("presents", []))
    absents_noms = noms(cr.get("absents", []))
    excused_noms = noms(cr.get("excused", []))

    # Date formatting
    date_cr = cr.get("date", "")
    try:
        date_fmt = datetime.fromisoformat(date_cr).strftime("%d/%m/%Y") if date_cr else "-"
    except Exception:
        date_fmt = date_cr or "-"

    type_reunion = cr.get("type_reunion", "reunion")
    type_label = {"ag": "Assemblee Generale", "ca": "Conseil d'Administration",
                  "bureau": "Reunion de Bureau", "reunion": "Reunion"}.get(type_reunion, type_reunion)

    # Votes: filter by same date as CR (or all if no date)
    votes = data.get("votes", [])
    votes_cr = [v for v in votes if v.get("date", "") == date_cr] if date_cr else []
    # If no votes match the exact date, include all votes of same type_vote
    if not votes_cr:
        votes_cr = [v for v in votes if v.get("type_vote", "") == type_reunion]

    # Build votes HTML
    votes_html = ""
    if votes_cr:
        votes_html = "<h2>Votes et Resolutions</h2>"
        for v in votes_cr:
            r = v.get("resultat", {})
            pour = r.get("pour", 0)
            contre = r.get("contre", 0)
            abstention = r.get("abstention", 0)
            statut = v.get("statut", "ouvert")
            _notes = v.get("notes", "")
            _notes_html = f'<p style="font-size:.85rem"><em>Notes: {_notes}</em></p>' if _notes else ""
            votes_html += f"""
<div style="border:1px solid #ddd;border-radius:8px;padding:1rem;margin-bottom:1rem">
<h3 style="margin-top:0;color:#0a335c">{v.get('titre', '')}</h3>
<p style="color:#666;font-size:.85rem">{v.get('description', '')}</p>
<table style="width:100%;border-collapse:collapse;margin:.5rem 0">
<tr><th style="background:#6c9186;color:#fff;padding:.4rem;text-align:center">Pour</th>
<th style="background:#dd83a9;color:#fff;padding:.4rem;text-align:center">Contre</th>
<th style="background:#999;color:#fff;padding:.4rem;text-align:center">Abstention</th></tr>
<tr><td style="text-align:center;padding:.4rem;border:1px solid #eee;font-size:1.2rem;font-weight:bold">{pour}</td>
<td style="text-align:center;padding:.4rem;border:1px solid #eee;font-size:1.2rem;font-weight:bold">{contre}</td>
<td style="text-align:center;padding:.4rem;border:1px solid #eee;font-size:1.2rem;font-weight:bold">{abstention}</td></tr>
</table>
<p style="font-size:.8rem;color:#999">Statut: {statut}</p>
{_notes_html}
</div>"""
    else:
        votes_html = '<h2>Votes et Resolutions</h2><p style="color:#999;font-style:italic">Aucun vote enregistre pour cette reunion.</p>'

    # Build discussions / decisions / actions sections
    def text_section(title, content):
        if not content:
            return ""
        # Escape HTML and convert newlines to <br>
        safe = content.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        return f"<h2>{title}</h2><div style=\"white-space:pre-wrap\">{safe}</div>"

    discussions_html = text_section("Discussions", cr.get("discussions", ""))
    decisions_html = text_section("Decisions", cr.get("decisions", ""))
    actions_html = text_section("Actions a suivre", cr.get("actions", ""))

    # Ordre du jour
    odj = cr.get("ordre_du_jour", "")
    odj_html = ""
    if odj:
        safe_odj = odj.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        odj_html = f"<h2>Ordre du Jour</h2><div style=\"white-space:pre-wrap\">{safe_odj}</div>"
    else:
        odj_html = "<h2>Ordre du Jour</h2><p style=\"color:#999;font-style:italic\">Non precise</p>"

    # Presents / Absents / Excuses lists
    def list_html(label, noms_list, color):
        if not noms_list:
            return f"<p><strong>{label}:</strong> <span style=\"color:#999;font-style:italic\">Aucun</span></p>"
        return f"<p><strong>{label}:</strong> {', '.join(noms_list)}</p>"

    html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8">
<title>PV - {cr.get('titre', 'Reunion')}</title>
<style>
@page {{ margin: 2cm }}
body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 2rem; color: #333 }}
h1 {{ color: #0a335c; text-align: center; border-bottom: 3px solid #0a335c; padding-bottom: .5rem; margin-bottom: .3rem }}
h2 {{ color: #0a335c; margin-top: 2rem; border-bottom: 1px solid #ddd; padding-bottom: .3rem }}
.subtitle {{ text-align: center; color: #666; margin-bottom: 2rem }}
.info-box {{ background: #f5f5f5; border-radius: 8px; padding: 1rem; margin: 1rem 0 }}
table {{ width: 100%; border-collapse: collapse }}
.signatures {{ margin-top: 3rem; display: flex; justify-content: space-between; gap: 2rem }}
.sig-block {{ flex: 1; text-align: center }}
.sig-line {{ border-top: 1px solid #333; margin-top: 3rem; padding-top: .3rem; font-size: .85rem; color: #666 }}
.footer {{ margin-top: 3rem; text-align: center; font-size: .8rem; color: #999 }}
@media print {{ body {{ max-width: none; padding: 0 }} .no-print {{ display: none }} }}
</style>
</head><body>
<h1>PROCES-VERBAL</h1>
<p class="subtitle">{type_label} — Les Ami(e)s de Romy</p>
<p style="text-align:center;color:#666">Association de prevention des violences infantiles</p>

<div class="info-box">
<p><strong>Date:</strong> {date_fmt}</p>
<p><strong>Type de reunion:</strong> {type_label}</p>
<p><strong>Titre:</strong> {cr.get('titre', '')}</p>
</div>

<h2>Presence</h2>
{list_html('Presents', presents_noms, '#6c9186')}
{list_html('Absents', absents_noms, '#dd83a9')}
{list_html('Excuses', excused_noms, '#999')}

{odj_html}

{discussions_html}

{decisions_html}

{votes_html}

{actions_html}

<h2>Signatures</h2>
<div class="signatures">
<div class="sig-block"><div class="sig-line">President(e) de seance</div></div>
<div class="sig-block"><div class="sig-line">Secretaire de seance</div></div>
</div>

<div class="footer">PV genere le {datetime.now().strftime("%d/%m/%Y")} par {user["username"]} — CRM Les Ami(e)s de Romy v{VERSION}</div>
<div class="no-print" style="text-align:center;margin-top:1rem"><button onclick="window.print()" style="padding:.5rem 1.5rem;font-size:1rem;cursor:pointer;border-radius:8px;border:1px solid #0a335c;background:#0a335c;color:#fff">Imprimer / PDF</button></div>
</body></html>"""

    return HTMLResponse(content=html)


# ============ F13: CAMPAGNES SMS / WHATSAPP ============
@app.get("/api/sms-campaigns")
async def list_sms_campaigns(user=Depends(require_referent)):
    data = load_db()
    return data.get("sms_campaigns", [])

@app.post("/api/sms-campaigns")
async def create_sms_campaign(req: SmsCampaignCreate, user=Depends(require_referent)):
    data = load_db()
    if "sms_campaigns" not in data:
        data["sms_campaigns"] = []
    # Resolve contact_ids to phone numbers
    contacts = data.get("contacts", [])
    if req.qualite:
        req.contact_ids = [c["id"] for c in contacts if c.get("qualite") == req.qualite and c.get("telephone")]
    numeros = list(req.destinataires)
    for c in contacts:
        if c["id"] in req.contact_ids and c.get("telephone"):
            numeros.append(c["telephone"])
    # Deduplicate
    numeros = list(set(numeros))
    camp = {
        "id": f"sms-{secrets.token_hex(8)}",
        "message": req.message,
        "destinataires": numeros,
        "nb_destinataires": len(numeros),
        "statut": "prete",  # prete, envoyee, echec
        "cree_par": user["username"], "cree_le": now_iso()
    }
    data["sms_campaigns"].append(camp)
    save_db(data); backup_db()
    return {"status": "ok", "campaign": camp}

@app.post("/api/sms-campaigns/{camp_id}/send")
async def send_sms_campaign(camp_id: str, user=Depends(require_referent)):
    data = load_db()
    for camp in data.get("sms_campaigns", []):
        if camp["id"] == camp_id:
            # TODO: integrer Twilio/OVH/WhatsApp Business API
            # Pour l instant on marque comme envoyee
            camp["statut"] = "envoyee"
            camp["envoye_le"] = now_iso()
            save_db(data); backup_db()
            return {"status": "ok", "message": f"Campagne prete ({camp['nb_destinataires']} destinataires). Integration SMS a configurer.", "campaign": camp}
    raise HTTPException(status_code=404, detail="Campagne non trouvee")

@app.delete("/api/sms-campaigns/{camp_id}")
async def delete_sms_campaign(camp_id: str, user=Depends(require_referent)):
    data = load_db()
    data["sms_campaigns"] = [c for c in data.get("sms_campaigns", []) if c["id"] != camp_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F14: RELATIONS PRESSE & MEDIAS ============
@app.get("/api/presse/contacts")
async def list_presse_contacts(user=Depends(require_referent)):
    data = load_db()
    return data.get("presse_contacts", [])

@app.post("/api/presse/contacts")
async def create_presse_contact(req: PresseContactCreate, user=Depends(require_referent)):
    data = load_db()
    if "presse_contacts" not in data:
        data["presse_contacts"] = []
    c = {"id": f"prs-{secrets.token_hex(8)}", **req.dict(), "cree_le": now_iso()}
    data["presse_contacts"].append(c)
    save_db(data); backup_db()
    return {"status": "ok", "contact": c}

@app.delete("/api/presse/contacts/{pc_id}")
async def delete_presse_contact(pc_id: str, user=Depends(require_referent)):
    data = load_db()
    data["presse_contacts"] = [c for c in data.get("presse_contacts", []) if c["id"] != pc_id]
    save_db(data); backup_db()
    return {"status": "ok"}

@app.get("/api/presse/releases")
async def list_presse_releases(user=Depends(require_referent)):
    data = load_db()
    return data.get("presse_releases", [])

@app.post("/api/presse/releases")
async def create_presse_release(req: PresseReleaseCreate, user=Depends(require_referent)):
    data = load_db()
    if "presse_releases" not in data:
        data["presse_releases"] = []
    r = {"id": f"rel-{secrets.token_hex(8)}", **req.dict(), "cree_par": user["username"], "cree_le": now_iso()}
    data["presse_releases"].append(r)
    save_db(data); backup_db()
    return {"status": "ok", "release": r}

@app.delete("/api/presse/releases/{rel_id}")
async def delete_presse_release(rel_id: str, user=Depends(require_referent)):
    data = load_db()
    data["presse_releases"] = [r for r in data.get("presse_releases", []) if r["id"] != rel_id]
    save_db(data); backup_db()
    return {"status": "ok"}

@app.get("/api/presse/couvertures")
async def list_presse_couvertures(user=Depends(require_auth)):
    data = load_db()
    return data.get("presse_couvertures", [])

@app.post("/api/presse/couvertures")
async def create_presse_couverture(req: PresseCouvertureCreate, user=Depends(require_referent)):
    data = load_db()
    if "presse_couvertures" not in data:
        data["presse_couvertures"] = []
    c = {"id": f"cov-{secrets.token_hex(8)}", **req.dict(), "cree_le": now_iso()}
    data["presse_couvertures"].append(c)
    save_db(data); backup_db()
    return {"status": "ok", "couverture": c}

@app.delete("/api/presse/couvertures/{cov_id}")
async def delete_presse_couverture(cov_id: str, user=Depends(require_referent)):
    data = load_db()
    data["presse_couvertures"] = [c for c in data.get("presse_couvertures", []) if c["id"] != cov_id]
    save_db(data); backup_db()
    return {"status": "ok"}


# ============ F15: ESPACE MEMBRE SELF-SERVICE ============
@app.get("/api/mon-profil")
async def get_mon_profil(user=Depends(require_auth)):
    data = load_db()
    cid = user.get("contact_id", "")
    if not cid:
        raise HTTPException(status_code=404, detail="Aucun contact associe a ce compte")
    for c in data.get("contacts", []):
        if c["id"] == cid:
            # Also return cotisation status
            cots = [ct for ct in data.get("cotisations", []) if ct.get("contact_id") == cid]
            events = data.get("evenements", [])
            my_events = []
            for e in events:
                pres = e.get("presence", {}).get(cid, "")
                if pres:
                    my_events.append({"id": e["id"], "titre": e.get("titre",""), "date": e.get("date",""), "presence": pres})
            return {
                "contact": c,
                "cotisations": cots,
                "events_inscrits": my_events
            }
    raise HTTPException(status_code=404, detail="Contact non trouve")

@app.put("/api/mon-profil")
async def update_mon_profil(req: MembreProfilUpdate, user=Depends(require_auth)):
    data = load_db()
    cid = user.get("contact_id", "")
    if not cid:
        raise HTTPException(status_code=404, detail="Aucun contact associe")
    for c in data.get("contacts", []):
        if c["id"] == cid:
            changes = []
            if req.telephone is not None and c.get("telephone") != req.telephone:
                changes.append(f"telephone: {c.get('telephone','')} -> {req.telephone}")
                c["telephone"] = req.telephone
            if req.email is not None and c.get("email") != req.email:
                changes.append(f"email modifie")
                c["email"] = req.email
            if req.ville is not None:
                c["ville"] = req.ville
                changes.append("ville modifiee")
            if req.code_postal is not None:
                c["code_postal"] = req.code_postal
                changes.append("code postal modifie")
            if changes:
                add_history(c, user["username"], "modification_self", " | ".join(changes))
                save_db(data); backup_db()
            return {"status": "ok", "contact": c}
    raise HTTPException(status_code=404, detail="Contact non trouve")


# ============ STORAGE (Nipogi via SSHFS) ENDPOINTS ============
import shutil as _shutil

def _validate_storage_filename(filename: str) -> str:
    """Sanitize a filename for storage: no path traversal, no separators."""
    if not filename or not filename.strip():
        raise HTTPException(status_code=400, detail="Nom de fichier requis")
    # Remove any path components
    safe = os.path.basename(filename)
    # Remove .. and separators
    safe = safe.replace("..", "").replace("/", "").replace("\\", "")
    # Allow only safe characters
    safe = re.sub(r'[^\w\.-]', '_', safe)
    if not safe:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")
    return safe

def _storage_available() -> bool:
    """Check if the Nipogi SSHFS mount is accessible."""
    return STORAGE_DIR.is_dir() and STORAGE_UPLOADS.is_dir()

def _file_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    types = {
        ".pdf": "PDF", ".jpg": "Image", ".jpeg": "Image", ".png": "Image",
        ".gif": "Image", ".webp": "Image", ".doc": "Word", ".docx": "Word",
        ".xls": "Excel", ".xlsx": "Excel", ".ppt": "PowerPoint", ".pptx": "PowerPoint",
        ".txt": "Texte", ".csv": "CSV", ".zip": "Archive", ".mp4": "Video",
        ".avi": "Video", ".mov": "Video", ".mp3": "Audio", ".wav": "Audio",
        ".odt": "OpenDocument", ".ods": "OpenDocument", ".odp": "OpenDocument",
        ".rtf": "RTF",
    }
    return types.get(ext, "Autre")

@app.post("/api/storage/upload")
async def storage_upload(request: Request, user=Depends(require_auth)):
    """Upload a file to the Nipogi storage (uploads/ subfolder)."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible (montage SSHFS inaccessible)")
    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Fichier requis")

    ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".doc", ".docx",
                          ".xls", ".xlsx", ".ppt", ".pptx", ".txt", ".csv", ".odt", ".ods",
                          ".odp", ".rtf", ".zip", ".mp4", ".avi", ".mov", ".mp3", ".wav"}
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Type de fichier non autorise: {ext}")

    MAX_FILE_SIZE = 100 * 1024 * 1024  # 100 Mo for Nipogi storage
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 100 Mo)")

    safe_name = _validate_storage_filename(file.filename)
    # Avoid overwriting: append suffix if file exists
    dest = STORAGE_UPLOADS / safe_name
    if dest.exists():
        stem = Path(safe_name).stem
        dest = STORAGE_UPLOADS / f"{stem}_{secrets.token_hex(4)}{ext}"
        safe_name = dest.name

    with open(dest, "wb") as f:
        f.write(content)

    return {
        "status": "ok",
        "file": {
            "filename": safe_name,
            "taille": len(content),
            "type": _file_type(safe_name),
            "date": now_iso(),
            "uploaded_by": user["username"],
            "subfolder": "uploads",
        }
    }

@app.get("/api/storage/files")
async def storage_files(subfolder: str = "uploads", user=Depends(require_auth)):
    """List files stored on the Nipogi storage."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible")

    valid_subfolders = {"uploads", "documents", "backups", "archives"}
    if subfolder not in valid_subfolders:
        raise HTTPException(status_code=400, detail="Sous-dossier invalide")

    target_dir = STORAGE_DIR / subfolder
    if not target_dir.is_dir():
        return []

    files = []
    for f in sorted(target_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            stat = f.stat()
            files.append({
                "filename": f.name,
                "taille": stat.st_size,
                "type": _file_type(f.name),
                "date": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "subfolder": subfolder,
            })
    return files

@app.get("/api/storage/download/{filename}")
async def storage_download(filename: str, subfolder: str = "uploads", user=Depends(require_auth)):
    """Download a file from the Nipogi storage."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible")

    valid_subfolders = {"uploads", "documents", "backups", "archives"}
    if subfolder not in valid_subfolders:
        raise HTTPException(status_code=400, detail="Sous-dossier invalide")

    safe_name = _validate_storage_filename(filename)
    filepath = STORAGE_DIR / subfolder / safe_name
    # Extra safety: ensure the resolved path is within the storage directory
    if not str(filepath.resolve()).startswith(str(STORAGE_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin invalide")
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable")

    return FileResponse(str(filepath), filename=safe_name)

@app.delete("/api/storage/{filename}")
async def storage_delete(filename: str, subfolder: str = "uploads", user=Depends(require_referent)):
    """Delete a file from the Nipogi storage. Referent/admin only."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible")

    valid_subfolders = {"uploads", "documents", "backups", "archives"}
    if subfolder not in valid_subfolders:
        raise HTTPException(status_code=400, detail="Sous-dossier invalide")

    safe_name = _validate_storage_filename(filename)
    filepath = STORAGE_DIR / subfolder / safe_name
    if not str(filepath.resolve()).startswith(str(STORAGE_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin invalide")
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable")

    filepath.unlink()
    return {"status": "ok", "deleted": safe_name}

class ArchiveRequest(BaseModel):
    filename: str
    source_subfolder: str = "uploads"

@app.post("/api/storage/archive")
async def storage_archive(req: ArchiveRequest, user=Depends(require_referent)):
    """Move a file from uploads/ (Pi) to archives/ (Nipogi) to free Pi space. Referent/admin only."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible")

    safe_name = _validate_storage_filename(req.filename)
    valid_sources = {"uploads", "documents"}
    if req.source_subfolder not in valid_sources:
        raise HTTPException(status_code=400, detail="Sous-dossier source invalide")

    source_path = STORAGE_DIR / req.source_subfolder / safe_name
    if not str(source_path.resolve()).startswith(str(STORAGE_DIR.resolve())):
        raise HTTPException(status_code=400, detail="Chemin invalide")
    if not source_path.exists() or not source_path.is_file():
        raise HTTPException(status_code=404, detail="Fichier source introuvable")

    dest_path = STORAGE_ARCHIVES / safe_name
    if dest_path.exists():
        stem = Path(safe_name).stem
        ext = Path(safe_name).suffix
        dest_path = STORAGE_ARCHIVES / f"{stem}_{secrets.token_hex(4)}{ext}"

    _shutil.move(str(source_path), str(dest_path))

    return {
        "status": "ok",
        "archived": dest_path.name,
        "source": req.source_subfolder,
        "destination": "archives",
        "archived_by": user["username"],
    }

@app.get("/api/storage/stats")
async def storage_stats(user=Depends(require_auth)):
    """Get storage usage stats for the Nipogi mount."""
    if not _storage_available():
        raise HTTPException(status_code=503, detail="Stockage Nipogi non disponible")

    # Try to get filesystem stats for the mount point
    try:
        stat = os.statvfs(str(STORAGE_DIR))
        total = stat.f_blocks * stat.f_frsize
        free = stat.f_bavail * stat.f_frsize
        used = total - free
    except Exception:
        total = free = used = 0

    # Per-subfolder usage
    subfolders = {}
    for name, path in [("uploads", STORAGE_UPLOADS), ("documents", STORAGE_DOCUMENTS),
                       ("backups", STORAGE_BACKUPS), ("archives", STORAGE_ARCHIVES)]:
        size = 0
        count = 0
        if path.is_dir():
            for f in path.rglob("*"):
                if f.is_file():
                    size += f.stat().st_size
                    count += 1
        subfolders[name] = {"taille": size, "nb_fichiers": count}

    total_used = sum(s["taille"] for s in subfolders.values())

    return {
        "available": _storage_available(),
        "total_bytes": total,
        "free_bytes": free,
        "used_bytes": used,
        "crm_storage_bytes": total_used,
        "subfolders": subfolders,
    }


# ============ F5-AG: GENERATEUR DOCUMENTS ASSEMBLEE GENERALE ============

ASSO_NAME = "Les Ami(e)s de Romy"
ASSO_SIEGE = "558 avenue du Vieux Mas, 13600 La Ciotat"

def _ag_auth(token: str):
    """Auth via token query param (for window.open)."""
    user = get_user_from_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="Token invalide")
    if user["role"] not in REFERENT_ROLES:
        raise HTTPException(status_code=403, detail="Acces referent requis")
    return user

def _ag_periode() -> str:
    """Retourne l'année associative (ex: '2025-2026'). Septembre→août."""
    now = datetime.now()
    if now.month >= 9:
        return f"{now.year}-{now.year + 1}"
    return f"{now.year - 1}-{now.year}"


def _ag_html_wrapper(title: str, body: str, user: dict, pdf_filename: str = "") -> str:
    """HTML wrapper with shared print-friendly CSS for AG documents.

    pdf_filename: nom propre du document (sans extension) qui devient document.title
    → c'est ce nom que le navigateur propose quand l'utilisateur fait « Enregistrer en PDF ».
    """
    pdf_name = pdf_filename or title
    return f"""<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8"><title>{pdf_name}</title>
<style>
body{{font-family:Arial,sans-serif;max-width:800px;margin:0 auto;padding:2rem;color:#333}}
h1{{color:#0a335c;text-align:center;border-bottom:3px solid #0a335c;padding-bottom:.5rem}}
h2{{color:#0a335c;margin-top:2rem;border-left:4px solid #6c9186;padding-left:.5rem}}
.asso{{text-align:center;color:#666;font-size:.9rem;margin-bottom:.5rem}}
.stat-box{{display:inline-block;width:45%;margin:1%;padding:1rem;background:#e8edf2;border-radius:10px;text-align:center}}
.stat-box .num{{font-size:2rem;font-weight:bold;color:#0a335c}}
.stat-box .lbl{{font-size:.85rem;color:#666}}
table{{width:100%;border-collapse:collapse;margin:1rem 0}}
th{{background:#0a335c;color:#fff;padding:.5rem;text-align:left}}
td{{padding:.5rem;border-bottom:1px solid #eee}}
tr:nth-child(even){{background:#f9f9f9}}
.checkbox{{width:20px;height:20px;border:2px solid #0a335c;display:inline-block;margin-right:.5rem;border-radius:3px}}
.signature{{margin-top:3rem;display:flex;justify-content:space-between}}
.signature div{{border-top:1px solid #333;padding-top:.3rem;width:40%;text-align:center;font-size:.85rem}}
.btn-print{{position:fixed;top:1rem;right:1rem;background:#0a335c;color:#fff;border:none;padding:10px 20px;border-radius:8px;cursor:pointer;font-size:1rem;z-index:999}}
.btn-print:hover{{background:#072645}}
.btn-pdf{{position:fixed;top:1rem;right:7.5rem;background:#6c9186;color:#fff;border:none;padding:10px 20px;border-radius:8px;cursor:pointer;font-size:1rem;z-index:999}}
.btn-pdf:hover{{background:#5a7d72}}
.footer{{margin-top:3rem;text-align:center;font-size:.8rem;color:#999}}
@media print{{.btn-print,.btn-pdf{{display:none}}}}
</style></head><body>
<button class="btn-pdf" onclick="downloadPDF()">&#128190; T&eacute;l&eacute;charger PDF</button>
<button class="btn-print" onclick="window.print()">&#128424; Imprimer</button>
<div class="asso">{ASSO_NAME} &mdash; Pr&eacute;vention des violences infantiles<br>{ASSO_SIEGE}</div>
{body}
<div class="footer">G&eacute;n&eacute;r&eacute; le {datetime.now().strftime("%d/%m/%Y")} par {user["username"]} — CRM v{VERSION}</div>
<script>
document.title = "{pdf_name}";
function downloadPDF() {{
  document.title = "{pdf_name}";
  window.print();
}}
</script>
</body></html>"""

@app.get("/api/ag/convocation")
async def ag_convocation(date_ag: str = "", lieu: str = "", ordre_du_jour: str = "", token: str = ""):
    """Génère une convocation d'AG en HTML."""
    user = _ag_auth(token)
    data = load_db()
    date_fmt = date_ag or "Date à définir"
    lieu_fmt = lieu or "Lieu à définir"
    odj_items = [l.strip() for l in ordre_du_jour.split("\n") if l.strip()] if ordre_du_jour else [
        "Ouverture de la séance",
        "Approbation du PV de la précédente AG",
        "Rapport moral",
        "Rapport financier",
        "Bilan d'activité",
        "Renouvellement du bureau",
        "Questions diverses",
    ]
    odj_html = "<ol>"
    for item in odj_items:
        odj_html += f"<li>{item}</li>"
    odj_html += "</ol>"

    # Count members à jour de cotisation
    contacts = data.get("contacts", [])
    cots = data.get("cotisations", [])
    annee_courante = str(datetime.now().year)
    ids_a_jour = {c.get("contact_id") for c in cots if c.get("annee") == annee_courante and c.get("statut") == "paye"}
    nb_a_jour = sum(1 for c in contacts if c["id"] in ids_a_jour)

    body = f"""
<h1>Convocation à l'Assemblée Générale</h1>
<p style="text-align:center;margin:1rem 0;font-size:1.1rem">Les membres de l'association <strong>{ASSO_NAME}</strong> sont convoqués à l'Assemblée Générale qui se tiendra :</p>
<table style="width:100%;margin:1.5rem 0">
<tr><th style="width:30%">Date</th><td>{date_fmt}</td></tr>
<tr><th>Lieu</th><td>{lieu_fmt}</td></tr>
<tr><th>Nombre de membres à jour de cotisation</th><td>{nb_a_jour}</td></tr>
</table>
<h2>Ordre du jour</h2>
{odj_html}
<p style="margin:1.5rem 0">Conformément aux statuts, tous les membres à jour de leur cotisation sont invités à participer. Les pouvoirs pour représentation doivent être donnés par écrit.</p>
<h2>Convocation</h2>
<p>La présente convocation est adressée à l'ensemble des {len(contacts)} contacts enregistrés, dont {nb_a_jour} membres à jour de cotisation pour l'année {annee_courante}.</p>
<div class="signature">
<div>Le Président / La Présidente</div>
<div>Le Secrétaire / La Secrétaire</div>
</div>
"""
    periode = _ag_periode()
    return HTMLResponse(content=_ag_html_wrapper("Convocation AG", body, user, f"Convocation_AG_{periode}"))


@app.get("/api/ag/feuille-presence")
async def ag_feuille_presence(token: str = ""):
    """Génère une feuille de présence d'AG en HTML — liste des adhérents à jour de cotisation."""
    user = _ag_auth(token)
    data = load_db()
    contacts = data.get("contacts", [])
    cots = data.get("cotisations", [])
    annee_courante = str(datetime.now().year)

    # Map contact_id → cotisation status for current year
    cot_by_id = {}
    for c in cots:
        if c.get("annee") == annee_courante:
            cot_by_id[c.get("contact_id")] = c

    rows = ""
    nb_a_jour = 0
    nb_non_a_jour = 0
    for c in contacts:
        cot = cot_by_id.get(c["id"])
        statut = cot.get("statut", "") if cot else "non cotisé"
        a_jour = "✓" if statut == "paye" else "✗"
        if statut == "paye":
            nb_a_jour += 1
        else:
            nb_non_a_jour += 1
        rows += f"<tr><td style='text-align:center'>{a_jour}</td><td>{c.get('prenom','')} {c.get('nom','')}</td><td>{c.get('qualite','')}</td><td>{c.get('email','')}</td><td style='text-align:center'>{statut}</td><td style='width:60px'>&nbsp;</td></tr>"

    body = f"""
<h1>Feuille de Présence — Assemblée Générale</h1>
<p style="text-align:center;margin:1rem 0">Date: ____ / ____ / {annee_courante} &nbsp;&nbsp;&mdash;&nbsp;&nbsp; Lieu: ____________________________</p>
<table>
<tr><th>À jour</th><th>Nom Prénom</th><th>Qualité</th><th>Email</th><th>Cotisation {annee_courante}</th><th>Signature</th></tr>
{rows}
</table>
<p style="margin-top:1rem"><strong>Membres à jour de cotisation: {nb_a_jour}</strong> / {len(contacts)} contacts enregistrés ({nb_non_a_jour} non à jour)</p>
<div class="signature">
<div>Le Secrétaire / La Secrétaire</div>
<div>Le Président / La Présidente</div>
</div>
"""
    periode = _ag_periode()
    return HTMLResponse(content=_ag_html_wrapper("Feuille de présence AG", body, user, f"Feuille_Presence_AG_{periode}"))


@app.get("/api/ag/rapport-moral")
async def ag_rapport_moral(token: str = ""):
    """Génère un rapport moral en HTML — basé sur les activités de l'association."""
    user = _ag_auth(token)
    data = load_db()
    contacts = data.get("contacts", [])
    events = data.get("evenements", [])
    accomps = data.get("accompagnements", [])
    cots = data.get("cotisations", [])
    dons = data.get("dons", [])
    annee = str(datetime.now().year)

    events_an = [e for e in events if e.get("date", "").startswith(annee)]
    accomps_an = [a for a in accomps if a.get("date_debut", "").startswith(annee)]
    cots_an = [c for c in cots if c.get("annee") == annee]
    dons_an = [d for d in dons if d.get("date", "").startswith(annee)]
    nb_a_jour = sum(1 for c in cots_an if c.get("statut") == "paye")

    body = f"""
<h1>Rapport Moral — {annee}</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">Présenté à l'Assemblée Générale par le Bureau de l'association {ASSO_NAME}</p>

<h2>Introduction</h2>
<p>L'association {ASSO_NAME}, dont le siège social est situé {ASSO_SIEGE}, a poursuivi en {annee} sa mission de prévention des violences infantiles. Ce rapport moral retrace les principales actions menées et les résultats obtenus.</p>

<h2>Vie associative</h2>
<p>Au cours de l'année {annee}, l'association compte <strong>{len(contacts)} contacts</strong> enregistrés, dont <strong>{nb_a_jour} membres à jour de cotisation</strong>. La vie associative s'est articulée autour des événements, des accompagnements et des actions de sensibilisation.</p>

<h2>Actions menées</h2>
<p><strong>{len(events_an)} événements</strong> ont été organisés en {annee} :</p>
<table>
<tr><th>Date</th><th>Titre</th><th>Lieu</th></tr>"""
    for e in events_an:
        body += f"<tr><td>{e.get('date','')}</td><td>{e.get('titre','')}</td><td>{e.get('lieu','')}</td></tr>"
    if not events_an:
        body += "<tr><td colspan='3' style='text-align:center;color:#999'>Aucun événement enregistré</td></tr>"
    body += f"""</table>

<h2>Accompagnements</h2>
<p><strong>{len(accomps_an)} accompagnements</strong> ont été suivis en {annee}.</p>

<h2>Ressources et financement</h2>
<p>Les ressources de l'association proviennent des cotisations ({nb_a_jour} cotisations payées pour {annee}) et des dons ({len(dons_an)} dons reçus). Ces ressources ont permis de financer les actions décrites ci-dessus.</p>

<h2>Perspectives</h2>
<p>L'association poursuit ses actions de prévention et de sensibilisation. Les perspectives pour l'année à venir concernent le développement de nouveaux partenariats, l'élargissement du réseau de bénévoles et le renforcement des actions d'accompagnement.</p>

<h2>Conclusion</h2>
<p>Le Bureau remercie l'ensemble des membres, bénévoles et partenaires pour leur engagement et leur soutien tout au long de l'année {annee}.</p>
<div class="signature">
<div>Le Président / La Présidente</div>
<div>Le Secrétaire / La Secrétaire</div>
</div>
"""
    return HTMLResponse(content=_ag_html_wrapper("Rapport Moral AG", body, user, f"Rapport_Moral_AG_{annee}"))


@app.get("/api/ag/rapport-financier")
async def ag_rapport_financier(token: str = ""):
    """Génère un rapport financier simplifié en HTML — depuis /api/stats + /api/dashboard."""
    user = _ag_auth(token)
    data = load_db()
    contacts = data.get("contacts", [])
    cots = data.get("cotisations", [])
    dons = data.get("dons", [])
    notes_frais = data.get("notes_frais", [])
    annee = str(datetime.now().year)

    cots_an = [c for c in cots if c.get("annee") == annee]
    dons_an = [d for d in dons if d.get("date", "").startswith(annee)]
    notes_an = [n for n in notes_frais if n.get("date", "").startswith(annee)]

    total_cots = sum(c.get("montant", 0) for c in cots_an if c.get("statut") == "paye")
    total_dons = sum(d.get("montant", 0) for d in dons_an)
    total_frais = sum(n.get("montant", 0) for n in notes_an)
    solde = total_cots + total_dons - total_frais

    # Cotisations par statut
    cots_payees = sum(1 for c in cots_an if c.get("statut") == "paye")
    cots_attente = sum(1 for c in cots_an if c.get("statut") == "en_attente")
    cots_impayees = sum(1 for c in cots_an if c.get("statut") == "impaye")

    body = f"""
<h1>Rapport Financier — {annee}</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">Présenté à l'Assemblée Générale par le Trésorier / La Trésorière</p>

<h2>Bilan simplifié</h2>
<div class="stat-box"><div class="num">{total_cots:.0f} €</div><div class="lbl">Cotisations perçues</div></div>
<div class="stat-box"><div class="num">{total_dons:.0f} €</div><div class="lbl">Dons reçus</div></div>
<div class="stat-box"><div class="num">{total_frais:.0f} €</div><div class="lbl">Notes de frais</div></div>
<div class="stat-box"><div class="num">{solde:.0f} €</div><div class="lbl">Solde (produits - charges)</div></div>

<h2>Détail des cotisations {annee}</h2>
<table>
<tr><th>Statut</th><th>Nombre</th><th>Montant</th></tr>
<tr><td>Payées</td><td style="text-align:center">{cots_payees}</td><td>{total_cots:.0f} €</td></tr>
<tr><td>En attente</td><td style="text-align:center">{cots_attente}</td><td>{sum(c.get('montant',0) for c in cots_an if c.get('statut')=='en_attente'):.0f} €</td></tr>
<tr><td>Impayées</td><td style="text-align:center">{cots_impayees}</td><td>{sum(c.get('montant',0) for c in cots_an if c.get('statut')=='impaye'):.0f} €</td></tr>
</table>

<h2>Dons de l'année</h2>
<table>
<tr><th>Date</th><th>Nom</th><th>Montant</th></tr>"""
    for d in dons_an:
        body += f"<tr><td>{d.get('date','')}</td><td>{d.get('nom','')}</td><td>{d.get('montant',0):.0f} €</td></tr>"
    if not dons_an:
        body += "<tr><td colspan='3' style='text-align:center;color:#999'>Aucun don enregistré</td></tr>"
    body += f"""</table>

<h2>Notes de frais de l'année</h2>
<table>
<tr><th>Date</th><th>Catégorie</th><th>Montant</th><th>Statut</th></tr>"""
    for n in notes_an:
        body += f"<tr><td>{n.get('date','')}</td><td>{n.get('categorie','')}</td><td>{n.get('montant',0):.0f} €</td><td>{n.get('statut','')}</td></tr>"
    if not notes_an:
        body += "<tr><td colspan='4' style='text-align:center;color:#999'>Aucune note de frais</td></tr>"
    body += f"""</table>

<h2>Synthèse</h2>
<table>
<tr><th>Produits</th><th>Charges</th><th>Solde</th></tr>
<tr><td style="text-align:center">{(total_cots + total_dons):.0f} €</td><td style="text-align:center">{total_frais:.0f} €</td><td style="text-align:center;font-weight:bold">{solde:.0f} €</td></tr>
</table>
<div class="signature">
<div>Le Trésorier / La Trésorière</div>
<div>Le Président / La Présidente</div>
</div>
"""
    return HTMLResponse(content=_ag_html_wrapper("Rapport Financier AG", body, user, f"Rapport_Financier_AG_{annee}"))


@app.get("/api/ag/bilan-activite")
async def ag_bilan_activite(token: str = "", annee: str = ""):
    """Génère un bilan d'activité en HTML — depuis événements et accompagnements."""
    user = _ag_auth(token)
    data = load_db()
    annee_cible = annee or str(datetime.now().year)
    events = data.get("evenements", [])
    accomps = data.get("accompagnements", [])

    events_an = [e for e in events if e.get("date", "").startswith(annee_cible)]
    accomps_an = [a for a in accomps if a.get("date_debut", "").startswith(annee_cible)]

    # Stats de présence
    presence_total = sum(sum(1 for v in e.get("presence", {}).values() if v == "present") for e in events_an)

    body = f"""
<h1>Bilan d'Activité — {annee_cible}</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">Présenté à l'Assemblée Générale</p>

<h2>Chiffres clés</h2>
<div class="stat-box"><div class="num">{len(events_an)}</div><div class="lbl">Événements organisés</div></div>
<div class="stat-box"><div class="num">{presence_total}</div><div class="lbl">Présences totales</div></div>
<div class="stat-box"><div class="num">{len(accomps_an)}</div><div class="lbl">Accompagnements</div></div>

<h2>Événements de l'année</h2>
<table>
<tr><th>Date</th><th>Titre</th><th>Lieu</th><th>Présences</th></tr>"""
    for e in events_an:
        pres = sum(1 for v in e.get("presence", {}).values() if v == "present")
        body += f"<tr><td>{e.get('date','')}</td><td>{e.get('titre','')}</td><td>{e.get('lieu','')}</td><td style='text-align:center'>{pres}</td></tr>"
    if not events_an:
        body += "<tr><td colspan='4' style='text-align:center;color:#999'>Aucun événement</td></tr>"
    body += f"""</table>

<h2>Accompagnements de l'année</h2>
<table>
<tr><th>Type</th><th>Statut</th><th>Priorité</th></tr>"""
    for a in accomps_an:
        body += f"<tr><td>{a.get('type_suivi','')}</td><td>{a.get('statut','')}</td><td>{a.get('priorite','')}</td></tr>"
    if not accomps_an:
        body += "<tr><td colspan='3' style='text-align:center;color:#999'>Aucun accompagnement</td></tr>"
    body += f"""</table>

<h2>Conclusion</h2>
<p>En {annee_cible}, l'association a organisé {len(events_an)} événements pour un total de {presence_total} présences, et a suivi {len(accomps_an)} accompagnements. L'engagement des bénévoles et la qualité des actions menées restent au cœur de la mission de prévention des violences infantiles.</p>
<div class="signature">
<div>Le Président / La Présidente</div>
<div>Le Secrétaire / La Secrétaire</div>
</div>
"""
    return HTMLResponse(content=_ag_html_wrapper("Bilan d'Activité AG", body, user, f"Bilan_Activite_AG_{annee_cible}"))


@app.get("/api/ag/bulletin-vote")
async def ag_bulletin_vote(titre: str = "", description: str = "", token: str = ""):
    """Génère un bulletin de vote en HTML pour une résolution."""
    user = _ag_auth(token)
    titre_res = titre or "Résolution à voter"
    desc = description or ""

    body = f"""
<h1>Bulletin de Vote</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">Assemblée Générale — {ASSO_NAME}</p>

<h2>Résolution</h2>
<p style="font-size:1.1rem;font-weight:bold;margin:1rem 0">{titre_res}</p>
{"<p>" + desc.replace(chr(10), "<br>") + "</p>" if desc else ""}

<h2>Vote</h2>
<p style="margin:1rem 0">Cochez une seule case :</p>
<table style="width:100%;border:2px solid #0a335c">
<tr><td style="padding:1rem;font-size:1.2rem"><span class="checkbox"></span> POUR</td></tr>
<tr><td style="padding:1rem;font-size:1.2rem"><span class="checkbox"></span> CONTRE</td></tr>
<tr><td style="padding:1rem;font-size:1.2rem"><span class="checkbox"></span> ABSTENTION</td></tr>
</table>

<p style="margin:1.5rem 0">Nom: _________________________________</p>
<p style="margin:.5rem 0">Prénom: _________________________________</p>
<p style="margin:.5rem 0">Signature:</p>
<p style="margin-top:2rem">_____________________________________</p>
<div class="footer" style="margin-top:3rem">Bulletin confidentiel — CRM v{VERSION}</div>
"""
    periode = _ag_periode()
    return HTMLResponse(content=_ag_html_wrapper("Bulletin de Vote", body, user, f"Bulletin_Vote_AG_{periode}"))


@app.get("/api/ag/pv")
async def ag_pv(date_ag: str = "", lieu: str = "", token: str = ""):
    """Génère un Procès-Verbal complet d'AG en HTML."""
    user = _ag_auth(token)
    data = load_db()
    contacts = data.get("contacts", [])
    cots = data.get("cotisations", [])
    votes = data.get("votes", [])
    annee = str(datetime.now().year)

    date_fmt = date_ag or datetime.now().strftime("%d/%m/%Y")
    lieu_fmt = lieu or ASSO_SIEGE

    # Members à jour
    cot_by_id = {}
    for c in cots:
        if c.get("annee") == annee:
            cot_by_id[c.get("contact_id")] = c
    presents_html = ""
    nb_a_jour = 0
    for c in contacts:
        cot = cot_by_id.get(c["id"])
        a_jour = cot and cot.get("statut") == "paye"
        if a_jour:
            nb_a_jour += 1
            presents_html += f"<tr><td style='text-align:center'>&#9745;</td><td>{c.get('prenom','')} {c.get('nom','')}</td><td>{c.get('qualite','')}</td></tr>"

    # Votes (scrutins)
    votes_html = ""
    for v in votes:
        r = v.get("resultat", {})
        pour = r.get("pour", 0)
        contre = r.get("contre", 0)
        abstention = r.get("abstention", 0)
        statut = v.get("statut", "ouvert")
        votes_html += f"<tr><td>{v.get('titre','')}</td><td>{v.get('type_vote','')}</td><td style='text-align:center'>{pour}</td><td style='text-align:center'>{contre}</td><td style='text-align:center'>{abstention}</td><td style='text-align:center'>{statut}</td></tr>"
    if not votes:
        votes_html = "<tr><td colspan='6' style='text-align:center;color:#999'>Aucun vote enregistré</td></tr>"

    body = f"""
<h1>Procès-Verbal de l'Assemblée Générale</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">{ASSO_NAME} — {ASSO_SIEGE}</p>
<table style="width:100%;margin:1rem 0">
<tr><th style="width:30%">Date</th><td>{date_fmt}</td></tr>
<tr><th>Lieu</th><td>{lieu_fmt}</td></tr>
<tr><th>Heure d'ouverture</th><td>____ h ____</td></tr>
<tr><th>Heure de clôture</th><td>____ h ____</td></tr>
<tr><th>Nombre de membres à jour de cotisation</th><td>{nb_a_jour}</td></tr>
<tr><th>Nombre de présents</th><td>____ (à remplir)</td></tr>
<tr><th>Nombre de pouvoirs</th><td>____</td></tr>
</table>

<h2>1. Ouverture de la séance</h2>
<p>L'Assemblée Générale s'est réunie le {date_fmt} à {lieu_fmt}. La séance a été ouverte par le Président / La Présidente.</p>

<h2>2. Présents</h2>
<table>
<tr><th>Présent</th><th>Nom Prénom</th><th>Qualité</th></tr>
{presents_html}
</table>

<h2>3. Approbation du PV de la précédente AG</h2>
<p>Le procès-verbal de la précédente Assemblée Générale est lu et approuvé à l'unanimité / à la majorité (rayer la mention inutile).</p>

<h2>4. Rapport moral</h2>
<p>Le rapport moral est présenté par le Président / La Présidente. Il est approuvé par l'Assemblée.</p>

<h2>5. Rapport financier</h2>
<p>Le rapport financier est présenté par le Trésorier / La Trésorière. Il est approuvé par l'Assemblée.</p>

<h2>6. Bilan d'activité</h2>
<p>Le bilan d'activité est présenté. Les actions menées au cours de l'année sont exposées et discutées.</p>

<h2>7. Délibérations et votes</h2>
<table>
<tr><th>Résolution</th><th>Type</th><th>Pour</th><th>Contre</th><th>Abstention</th><th>Statut</th></tr>
{votes_html}
</table>

<h2>8. Questions diverses</h2>
<p>Diverses questions ont été abordées par les membres présents.</p>

<h2>9. Clôture de la séance</h2>
<p>La séance est levée à ____ h ____.</p>

<div class="signature">
<div>Le Président / La Présidente</div>
<div>Le Secrétaire / La Secrétaire</div>
</div>
"""
    periode = _ag_periode()
    return HTMLResponse(content=_ag_html_wrapper("Procès-Verbal AG", body, user, f"Proces-Verbal_AG_{periode}"))


@app.get("/api/ag/checklist")
async def ag_checklist(token: str = ""):
    """Génère une checklist pré-AG basée sur les chiffres réels du CRM pour l'année écoulée (septembre à août)."""
    user = _ag_auth(token)
    data = load_db()

    # Année associative : septembre N-1 à août N
    # Exemple : en août 2026, l'année écoulée = sept 2025 à août 2026
    # En septembre 2026, l'année écoulée = sept 2025 à août 2026 aussi
    now = datetime.now()
    if now.month >= 9:
        annee_debut = now.year
        annee_fin = now.year + 1
    else:
        annee_debut = now.year - 1
        annee_fin = now.year
    periode = f"{annee_debut}-{annee_fin}"
    date_debut = f"{annee_debut}-09-01"
    date_fin = f"{annee_fin}-08-31"
    annee_cot = str(annee_fin)  # année cotisation = année de fin

    contacts = data.get("contacts", [])
    cots = data.get("cotisations", [])
    events = data.get("evenements", [])
    dons = data.get("dons", [])
    notes_frais = data.get("notes_frais", [])
    votes = data.get("votes", [])
    accompagnements = data.get("accompagnements", [])
    documents = data.get("documents", [])

    # Chiffres réels de l'année écoulée
    nb_contacts_total = len(contacts)
    nb_cots_annee = len([c for c in cots if c.get("annee") == annee_cot])
    nb_cots_payees = len([c for c in cots if c.get("annee") == annee_cot and c.get("statut") == "paye"])
    nb_cots_impayees = len([c for c in cots if c.get("annee") == annee_cot and c.get("statut") != "paye"])
    montant_cots = sum(c.get("montant", 0) for c in cots if c.get("annee") == annee_cot and c.get("statut") == "paye")

    # Événements de l'année écoulée
    events_annee = [e for e in events if date_debut <= e.get("date", "") <= date_fin]
    nb_events = len(events_annee)
    events_a_venir = [e for e in events if e.get("date", "") >= now.strftime("%Y-%m-%d")]

    # Interventions de sensibilisation
    interventions = [e for e in events_annee if e.get("type", "") == "intervention"]
    nb_interventions = len(interventions)
    nb_participants_total = sum(len(e.get("participants", [])) for e in interventions)

    # Dons de l'année
    dons_annee = [d for d in dons if date_debut <= d.get("date", "") <= date_fin]
    nb_dons = len(dons_annee)
    montant_dons = sum(d.get("montant", 0) for d in dons_annee)

    # Notes de frais
    frais_annee = [f for f in notes_frais if date_debut <= f.get("date", "") <= date_fin]
    nb_frais = len(frais_annee)
    montant_frais = sum(f.get("montant", 0) for f in frais_annee)

    # Accompagnements actifs
    accomp_actifs = [a for a in accompagnements if a.get("statut", "") == "actif"]

    # Votes enregistrés
    nb_votes = len(votes)

    # Documents dans le CRM
    nb_docs = len(documents)

    # Items de la checklist avec chiffres réels
    items = [
        ("Définir la date et le lieu de l'AG", "Convocations", ""),
        ("Réserver la salle", "Logistique", ""),
        ("Préparer les convocations (courrier/email)", "Convocations", ""),
        ("Envoyer les convocations (15 jours minimum avant l'AG)", "Convocations", ""),
        (f"Vérifier les cotisations {annee_cot} : {nb_cots_payees} payées, {nb_cots_impayees} impayées sur {nb_cots_annee} attendues", "Adhésions", f"{nb_cots_payees}/{nb_contacts_total}"),
        (f"Relancer les {nb_cots_impayees} adhérents en retard de cotisation", "Adhésions", f"{nb_cots_impayees} relances"),
        (f"Préparer le rapport moral (exercice {periode})", "Documents", ""),
        (f"Préparer le rapport financier : {montant_cots}€ cotisations + {montant_dons}€ dons + {montant_frais}€ frais", "Documents", f"{montant_cots + montant_dons}€ recettes"),
        (f"Préparer le bilan d'activité : {nb_interventions} interventions, {nb_participants_total} participants", "Documents", f"{nb_interventions} interv."),
        (f"Préparer le bilan salons/colloques : {nb_events - nb_interventions} événements autres", "Documents", ""),
        ("Préparer les feuilles de présence (adhérents à jour)", "Documents", f"{nb_cots_payees} à jour"),
        (f"Préparer les bulletins de vote : {nb_votes} résolution(s) enregistrée(s)", "Documents", f"{nb_votes} résolutions"),
        ("Préparer le PV (modèle)", "Documents", ""),
        ("Vérifier le matériel (urne, bulletins, procuration)", "Logistique", ""),
        ("Préparer le powerpoint / support de présentation", "Logistique", ""),
        ("Confirmer les intervenants (Président, Trésorier, Secrétaire)", "Organisation", ""),
        (f"Rappeler la date aux {nb_cots_payees} membres à jour (J-2)", "Convocations", ""),
        (f"Vérifier les {len(accomp_actifs)} accompagnements actifs à présenter", "Organisation", f"{len(accomp_actifs)} actifs"),
        (f"Archiver les {nb_docs} documents du CRM pour l'exercice", "Clôture", f"{nb_docs} docs"),
        ("Clôturer l'exercice comptable dans le CRM", "Clôture", ""),
    ]

    rows = ""
    for i, (item, cat, chiffre) in enumerate(items, 1):
        chiffre_html = f"<span style='background:#e8edf2;color:#0a335c;padding:2px 8px;border-radius:10px;font-size:.8rem;font-weight:600'>{chiffre}</span>" if chiffre else ""
        rows += f"<tr><td style='text-align:center'><span class='checkbox'></span></td><td style='text-align:center;font-weight:600'>{i}</td><td>{item}</td><td style='text-align:center'>{cat}</td><td style='text-align:center'>{chiffre_html}</td></tr>"

    # Tableau récapitulatif des chiffres de l'année
    recap_rows = ""
    recap_data = [
        ("Période de l'exercice", periode),
        ("Total contacts", str(nb_contacts_total)),
        ("Cotisations payées", f"{nb_cots_payees} ({montant_cots}€)"),
        ("Cotisations impayées", f"{nb_cots_impayees}"),
        ("Dons perçus", f"{nb_dons} ({montant_dons}€)"),
        ("Notes de frais", f"{nb_frais} ({montant_frais}€)"),
        ("Interventions de sensibilisation", f"{nb_interventions}"),
        ("Participants touchés", f"{nb_participants_total}"),
        ("Événements totaux", f"{nb_events}"),
        ("Accompagnements actifs", f"{len(accomp_actifs)}"),
        ("Votes enregistrés", f"{nb_votes}"),
        ("Documents dans le CRM", f"{nb_docs}"),
    ]
    for label, valeur in recap_data:
        recap_rows += f"<tr><td style='padding:8px;border:1px solid #ddd'>{label}</td><td style='padding:8px;border:1px solid #ddd;font-weight:600;color:#0a335c;text-align:right'>{valeur}</td></tr>"

    body = f"""
<h1>Checklist pré-Assemblée Générale — Exercice {periode}</h1>
<p style="text-align:center;font-style:italic;margin:1rem 0">{ASSO_NAME}</p>

<div style="background:#e8edf2;border-radius:12px;padding:1rem 1.5rem;margin:1.5rem 0">
<h2 style="color:#0a335c;margin-top:0">Récapitulatif de l'année écoulée ({periode})</h2>
<table style="width:100%;border-collapse:collapse">
{recap_rows}
</table>
</div>

<h2>Checklist de préparation</h2>
<table>
<tr><th>Fait</th><th>N°</th><th>Tâche</th><th>Catégorie</th><th>Chiffre clé</th></tr>
{rows}
</table>

<h2>Événements à venir</h2>
<table>
<tr><th>Date</th><th>Titre</th><th>Lieu</th></tr>"""
    for e in events_a_venir[:5]:
        body += f"<tr><td>{e.get('date','')}</td><td>{e.get('titre','')}</td><td>{e.get('lieu','')}</td></tr>"
    if not events_a_venir:
        body += "<tr><td colspan='3' style='text-align:center;color:#999'>Aucun événement à venir</td></tr>"
    body += """
</table>
<p style="margin-top:2rem;font-size:.9rem;color:#666">Cochez chaque élément au fur et à mesure de sa réalisation. Les chiffres sont extraits automatiquement du CRM pour l'exercice """ + periode + """ (septembre à août).</p>
"""
    return HTMLResponse(content=_ag_html_wrapper("Checklist pré-AG", body, user, f"Checklist_pre-AG_{periode}"))


# ============ SUBVENTIONS (v1.34) ============
@app.get("/api/subventions")
async def list_subventions(statut: str = "", user=Depends(require_referent)):
    data = load_db()
    subs = data.get("subventions", [])
    if statut:
        subs = [s for s in subs if s.get("statut") == statut]
    return subs

@app.post("/api/subventions")
async def create_subvention(req: SubventionCreate, user=Depends(require_referent)):
    data = load_db()
    if "subventions" not in data:
        data["subventions"] = []
    sub = {
        "id": f"sub-{secrets.token_hex(8)}",
        "organisme": req.organisme,
        "intitule": req.intitule,
        "montant_demande": req.montant_demande,
        "montant_accorde": req.montant_accorde,
        "date_demande": req.date_demande,
        "date_reponse": req.date_reponse,
        "statut": req.statut,
        "echeance": req.echeance,
        "documents_requis": req.documents_requis,
        "documents_remis": req.documents_remis,
        "notes": req.notes,
        "cree_par": user["username"],
        "cree_le": now_iso()
    }
    data["subventions"].append(sub)
    save_db(data)
    backup_db()
    return {"status": "ok", "subvention": sub}

@app.put("/api/subventions/{sub_id}")
async def update_subvention(sub_id: str, req: SubventionUpdate, user=Depends(require_referent)):
    data = load_db()
    for s in data.get("subventions", []):
        if s["id"] == sub_id:
            if req.organisme is not None: s["organisme"] = req.organisme
            if req.intitule is not None: s["intitule"] = req.intitule
            if req.montant_demande is not None: s["montant_demande"] = req.montant_demande
            if req.montant_accorde is not None: s["montant_accorde"] = req.montant_accorde
            if req.date_demande is not None: s["date_demande"] = req.date_demande
            if req.date_reponse is not None: s["date_reponse"] = req.date_reponse
            if req.statut is not None: s["statut"] = req.statut
            if req.echeance is not None: s["echeance"] = req.echeance
            if req.documents_requis is not None: s["documents_requis"] = req.documents_requis
            if req.documents_remis is not None: s["documents_remis"] = req.documents_remis
            if req.notes is not None: s["notes"] = req.notes
            save_db(data); backup_db()
            return {"status": "ok", "subvention": s}
    raise HTTPException(status_code=404, detail="Subvention non trouvee")

@app.delete("/api/subventions/{sub_id}")
async def delete_subvention(sub_id: str, user=Depends(require_referent)):
    data = load_db()
    data["subventions"] = [s for s in data.get("subventions", []) if s["id"] != sub_id]
    save_db(data); backup_db()
    return {"status": "ok"}

@app.get("/api/subventions/total")
async def total_subventions(user=Depends(require_referent)):
    data = load_db()
    subs = data.get("subventions", [])
    return {
        "total_demande": sum(s.get("montant_demande", 0) for s in subs),
        "total_accorde": sum(s.get("montant_accorde", 0) for s in subs),
        "brouillon": len([s for s in subs if s.get("statut") == "brouillon"]),
        "depose": len([s for s in subs if s.get("statut") == "depose"]),
        "accepte": len([s for s in subs if s.get("statut") == "accepte"]),
        "refuse": len([s for s in subs if s.get("statut") == "refuse"]),
        "count": len(subs)
    }


# ============ STARTUP ============
@app.on_event("startup")
async def startup():
    migrate_plaintext_passwords()
    validate_db_integrity()
    # Demarrer le thread de rappel automatique J-2
    t = threading.Thread(target=_reminder_check_loop, daemon=True)
    t.start()
    print(f"[CRM Romy] v{VERSION} demarre sur {HOST}:{PORT}")
    print(f"[CRM Romy] Base: {DB_PATH}")
    print(f"[CRM Romy] Contacts: {len(load_db().get('contacts', []))}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=HOST, port=PORT)